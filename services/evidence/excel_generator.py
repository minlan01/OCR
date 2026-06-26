"""
Excel 费用生成器 — 生成赔偿费用总表和医疗费用汇总表 Excel
========================================================
表1: 赔偿费用清单 — 标准8大赔偿项目 + 计算依据 + 金额
表2: 医疗费用汇总表 — 按医院分组、逐条明细（住院/门诊、报销/自费）
新增: generate_fee_excel_zip — 10项赔偿明细 ZIP 包
"""
from __future__ import annotations

import io
import re
import uuid
import zipfile
from typing import Any

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side, PatternFill
from openpyxl.utils import get_column_letter

MINIO_BUCKET = "scan-result"


def _clean_text(value: str) -> str:
    """清理文本字段：去除LLM返回的JSON标记、代码块标记等噪声"""
    if not value or not isinstance(value, str):
        return ""
    import re
    text = value.strip()
    # 去除代码块标记
    text = re.sub(r'^```(?:json|JSON)?\s*', '', text)
    text = re.sub(r'\s*```$', '', text)
    # 去除JSON对象/数组标记
    if text.startswith('{') and text.endswith('}'):
        return ""  # 整个值是JSON对象，不是有效文本
    if text.startswith('[') and text.endswith(']'):
        return ""  # 整个值是JSON数组
    # 去除多余的换行和空格
    text = re.sub(r'\n{2,}', '\n', text)
    text = text.strip()
    return text


def _get_case_data(case_id: str) -> tuple[dict, dict]:
    """获取案件的清单数据和分析结果"""
    from db.models_evidence import EvidenceCase
    from db.session import get_session_factory, run_in_worker

    async def _fetch():
        from sqlalchemy import select

        case_uuid = uuid.UUID(case_id)
        async with get_session_factory()() as db:
            stmt = select(EvidenceCase).where(EvidenceCase.id == case_uuid)
            result = await db.execute(stmt)
            case = result.scalar_one_or_none()
            if not case:
                raise ValueError(f"Case not found: {case_id}")
            return case.catalog_data or {}, case.analysis_result or {}

    return run_in_worker(_fetch())


def _upload_to_minio(case_id: str, excel_bytes: bytes, filename: str) -> str:
    """上传 Excel 到 MinIO"""
    from services.storage.minio_client import minio_client

    minio_key = f"evidence/{case_id}/{uuid.uuid4()}_{filename}"
    minio_client.upload_bytes(
        bucket=MINIO_BUCKET,
        object_key=minio_key,
        data=excel_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    return minio_key


# ─── 样式定义（统一仿宋字体） ─────────────────────────────────────────────

_FANGSONG_TITLE = Font(name="仿宋", size=16, bold=True)
_FANGSONG_HEADER = Font(name="仿宋", size=12, bold=True)
_FANGSONG_CELL = Font(name="仿宋", size=12)
_FANGSONG_BOLD = Font(name="仿宋", size=12, bold=True)

# 保留旧常量名做向后兼容（引用旧赔偿费用清单等）
_HEADER_FONT = Font(name="仿宋", size=12, bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_CELL_FONT = _FANGSONG_CELL
_BOLD_FONT = _FANGSONG_BOLD
_MONEY_FORMAT = '#,##0.00'
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
_LEFT_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_RIGHT_ALIGN = Alignment(horizontal="center", vertical="center")
_SUBTITLE_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
_SUBTOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_TOTAL_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")


def _apply_header_style(ws, row: int, col_count: int) -> None:
    """应用表头样式 — 仿宋12号加粗，居中"""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _FANGSONG_HEADER
        cell.alignment = _CENTER_ALIGN
        cell.border = _THIN_BORDER


def _apply_cell_style(ws, row: int, col_count: int, money_cols: set[int] | None = None) -> None:
    """应用单元格样式 — 仿宋12号，上下左右居中"""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _FANGSONG_CELL
        cell.border = _THIN_BORDER
        cell.alignment = _CENTER_ALIGN
        if money_cols and col in money_cols:
            cell.number_format = _MONEY_FORMAT


def _apply_row_border(ws, row: int, col_count: int) -> None:
    """给一行所有单元格加边框"""
    for col in range(1, col_count + 1):
        ws.cell(row=row, column=col).border = _THIN_BORDER


# ═══════════════════════════════════════════════════════════════════════════════
# 标准 8 大赔偿项目定义
# ═══════════════════════════════════════════════════════════════════════════════

# 每项: (显示名称, 计算依据模板, fee_summary 中的关键词匹配列表)
_STANDARD_COMPENSATION_ITEMS = [
    ("医疗费", "凭票据实报实销", ["医疗费", "医药费", "门诊费", "住院医疗"]),
    ("护理费", "住院期间护理费 = 护理标准 × 住院天数", ["护理费"]),
    ("住院伙食补助费", "住院伙食补助标准 × 住院天数", ["住院伙食补助费", "伙食补助费", "住院伙食费"]),
    ("营养费", "营养费标准 × 营养期天数（参照鉴定意见）", ["营养费"]),
    ("残疾赔偿金", "受诉法院所在地上年度城镇居民人均可支配收入 × 20年 × 伤残系数", ["残疾赔偿金", "伤残赔偿金"]),
    ("死亡赔偿金", "受诉法院所在地上年度城镇居民人均可支配收入 × 20年", ["死亡赔偿金"]),
    ("丧葬费", "受诉法院所在地上年度职工月平均工资 × 6个月", ["丧葬费"]),
    ("交通费", "凭票据实报实销", ["交通费"]),
    ("精神损害抚慰金", "根据伤残等级/死亡后果酌定", ["精神损害抚慰金", "精神抚慰金", "精神损害"]),
    ("鉴定费", "凭票据实报实销", ["鉴定费"]),
    ("误工费", "误工收入 × 误工天数（参照鉴定意见）", ["误工费"]),
    ("后续治疗费", "参照鉴定意见或后续实际发生费用", ["后续治疗费"]),
    ("被扶养人生活费", "受诉法院所在地上年度城镇居民人均消费支出 × 扶养年限", ["被扶养人生活费"]),
    ("残疾辅助器具费", "凭票据或参照普通适用器具标准", ["残疾辅助器具费", "辅助器具费"]),
    ("住宿费", "凭票据实报实销", ["住宿费"]),
    ("其他费用", "其他合理费用", ["其他费用", "其他"]),
]


def _match_fee_amount(fee_summary: dict, keywords: list[str]) -> float:
    """从 fee_summary 中按关键词匹配金额（最佳匹配而非首次匹配）

    匹配优先级：
    1. 完全匹配（fee_type == keyword）
    2. fee_type 包含 keyword
    3. keyword 包含 fee_type
    """
    best_match = 0.0
    best_score = 0

    for keyword in keywords:
        for fee_type, amount in fee_summary.items():
            if not isinstance(amount, (int, float)) or amount <= 0:
                continue

            # 计算匹配分数
            score = 0
            if fee_type == keyword:
                score = 100  # 完全匹配
            elif keyword in fee_type:
                score = 50   # fee_type 包含 keyword
            elif fee_type in keyword:
                score = 25   # keyword 包含 fee_type

            if score > best_score:
                best_score = score
                best_match = amount

    return best_match


# ═══════════════════════════════════════════════════════════════════════════════
# 旧版兼容函数（直接调 DB + 上传 MinIO）
# ═══════════════════════════════════════════════════════════════════════════════

def generate_compensation_summary(case_id: str) -> str:
    """生成赔偿费用总表 Excel → 返回 MinIO key"""
    catalog_data, analysis_result = _get_case_data(case_id)
    excel_bytes = generate_compensation_inline_data(catalog_data, analysis_result)
    if not excel_bytes:
        raise ValueError("Failed to generate compensation summary")
    minio_key = _upload_to_minio(case_id, excel_bytes, "赔偿费用清单.xlsx")
    logger.info(f"Compensation summary generated: case={case_id} key={minio_key}")
    return minio_key


def generate_fee_type_detail(case_id: str, fee_type: str) -> str:
    """生成单项费用明细 Excel（如医疗费.xlsx） → 返回 MinIO key"""
    catalog_data, analysis_result = _get_case_data(case_id)
    details = generate_fee_details_inline_data(catalog_data, analysis_result)
    if fee_type in details:
        minio_key = _upload_to_minio(case_id, details[fee_type], f"{fee_type}.xlsx")
        logger.info(f"Fee detail generated: case={case_id} type={fee_type} key={minio_key}")
        return minio_key
    raise ValueError(f"No fee detail for type: {fee_type}")


def generate_all_fee_details(case_id: str) -> dict[str, str]:
    """生成所有费用类型的独立 Excel → 返回 {fee_type: minio_key}"""
    catalog_data, analysis_result = _get_case_data(case_id)
    details = generate_fee_details_inline_data(catalog_data, analysis_result)
    results: dict[str, str] = {}
    for fee_type, excel_bytes in details.items():
        try:
            minio_key = _upload_to_minio(case_id, excel_bytes, f"{fee_type}.xlsx")
            results[fee_type] = minio_key
        except Exception as e:
            logger.error(f"Failed to upload fee detail for {fee_type}: {e}")
    return results


# ═══════════════════════════════════════════════════════════════════════════════
# 纯数据驱动生成函数（不调数据库，供路由端直接使用）
# ═══════════════════════════════════════════════════════════════════════════════

def generate_compensation_inline_data(catalog_data: dict, analysis_result: dict) -> bytes | None:
    """赔偿费用清单 Excel — 参考赔偿费用清单（李明凤）.xls 格式

    格式特点（精确匹配参考模板）：
    - 标题行（行1）：合并 A1:D1，"赔偿费用清单"，仿宋16号
    - 表头行（行2）：序号 | 项目 | 计算依据 | 金额（元），仿宋12号
    - 数据行：7大标准赔偿项目（医疗费/护理费/住院伙食补助费/营养费/交通住宿费/鉴定费/精神损害抚慰金）
    - 计算依据为详细法律计算说明（含住院天数、标准、公式）
    - 护理费支持多行计算依据（合并序号列和项目列）
    - 总计行：合并A:B写"总计"，合计金额
    - 备注行：合并A:D，鉴定前备注说明
    - 全部仿宋12号，上下左右居中
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "赔偿费用清单"

    case_name = analysis_result.get("case_name", "")
    case_type = analysis_result.get("case_type", "injury")

    # ── 提取住院天数信息 ──
    admission_date = ""
    discharge_date = ""
    stay_days = 0
    treatment = analysis_result.get("treatment", {}) or {}
    admission_date = treatment.get("admission_date", "") or ""
    discharge_date = treatment.get("discharge_date", "") or ""

    # 从 catalog_data 的 treatment 数据中找住院天数
    if not stay_days:
        for group in catalog_data.get("groups", []):
            for item in group.get("items", []):
                t = item.get("treatment", {}) or {}
                ad = t.get("admission_date", "")
                dd = t.get("discharge_date", "")
                if ad and dd:
                    try:
                        from datetime import datetime as _dt
                        d1 = _dt.strptime(ad[:10], "%Y-%m-%d")
                        d2 = _dt.strptime(dd[:10], "%Y-%m-%d")
                        stay_days = (d2 - d1).days
                        admission_date = ad[:10]
                        discharge_date = dd[:10]
                        break
                    except (ValueError, IndexError):
                        pass
            if stay_days:
                break

    # 从 fee_summary 获取金额
    fee_summary = catalog_data.get("fee_summary", {})

    # ── 提取各赔偿项金额 ──
    medical_amount = _match_fee_amount(fee_summary, ["医疗费", "医药费", "门诊费", "住院医疗"]) or 0
    nursing_amount = _match_fee_amount(fee_summary, ["护理费"]) or 0
    nursing_supplies = _match_fee_amount(fee_summary, ["护理用品", "医护用品"]) or 0
    food_amount = _match_fee_amount(fee_summary, ["住院伙食补助费", "伙食补助费", "住院伙食费"]) or 0
    nutrition_amount = _match_fee_amount(fee_summary, ["营养费"]) or 0
    transport_amount = _match_fee_amount(fee_summary, ["交通费", "住宿费", "交通住宿"]) or 0
    appraisal_amount = _match_fee_amount(fee_summary, ["鉴定费"]) or 0
    spiritual_amount = _match_fee_amount(fee_summary, ["精神损害抚慰金", "精神抚慰金", "精神损害"]) or 0

    # 从 compensation_data 中取手动编辑的金额（优先）
    comp_items = analysis_result.get("compensation_items", [])
    if comp_items:
        comp_map = {item.get("fee_type", ""): item for item in comp_items}
        for ft, default in [
            ("medical_fee", medical_amount), ("nursing_fee", nursing_amount),
            ("food_subsidy", food_amount), ("nutrition_fee", nutrition_amount),
            ("transport_fee", transport_amount), ("appraisal_fee", appraisal_amount),
            ("spiritual_damage", spiritual_amount),
        ]:
            item = comp_map.get(ft)
            if item:
                ma = item.get("manual_amount")
                if ma is not None:
                    if ft == "medical_fee":
                        medical_amount = _to_float(ma)
                    elif ft == "nursing_fee":
                        nursing_amount = _to_float(ma)
                    elif ft == "food_subsidy":
                        food_amount = _to_float(ma)
                    elif ft == "nutrition_fee":
                        nutrition_amount = _to_float(ma)
                    elif ft == "transport_fee":
                        transport_amount = _to_float(ma)
                    elif ft == "appraisal_fee":
                        appraisal_amount = _to_float(ma)
                    elif ft == "spiritual_damage":
                        spiritual_amount = _to_float(ma)

    COL_COUNT = 4

    # ── 行1：标题行（合并 A1:D1）──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=COL_COUNT)
    title_cell = ws.cell(row=1, column=1, value="赔偿费用清单")
    title_cell.font = _FANGSONG_TITLE
    title_cell.alignment = _CENTER_ALIGN

    # ── 行2：表头行 ──
    headers = ["序号", "项目", "计算依据", "金额（元）"]
    HEADER_ROW = 2
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=header)
        cell.font = _FANGSONG_HEADER
        cell.alignment = _CENTER_ALIGN
        cell.border = _THIN_BORDER

    # ── 数据行 ──
    row_idx = HEADER_ROW + 1
    seq = 1
    total_amount = 0.0

    def _write_item(item_seq, name, basis, amount, is_multi_row=False):
        """写一行赔偿项目。如果 is_multi_row=True，后续行会合并序号列和项目列。"""
        nonlocal row_idx, total_amount
        start_row = row_idx
        ws.cell(row=row_idx, column=1, value=item_seq)
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value=basis)
        ws.cell(row=row_idx, column=4, value=amount if amount else "")
        for col in range(1, COL_COUNT + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = _FANGSONG_CELL
            cell.border = _THIN_BORDER
            cell.alignment = _CENTER_ALIGN
            if col == 4 and amount and isinstance(amount, (int, float)):
                cell.number_format = _MONEY_FORMAT
        if amount and isinstance(amount, (int, float)) and amount > 0:
            total_amount += amount
        row_idx += 1

    def _write_continuation(text):
        """写计算依据续行（合并序号列和项目列，只有计算依据有值）。"""
        nonlocal row_idx
        ws.cell(row=row_idx, column=3, value=text)
        for col in range(1, COL_COUNT + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = _FANGSONG_CELL
            cell.border = _THIN_BORDER
            cell.alignment = _CENTER_ALIGN
        row_idx += 1

    def _merge_seq_project(start_row_idx, end_row_idx):
        """合并序号列(A)和项目列(B)的指定行范围。"""
        if end_row_idx > start_row_idx:
            ws.merge_cells(start_row=start_row_idx, start_column=1, end_row=end_row_idx, end_column=1)
            ws.merge_cells(start_row=start_row_idx, start_column=2, end_row=end_row_idx, end_column=2)

    # 1. 医疗费
    medical_basis = ""
    if medical_amount > 0:
        medical_basis = f"原告方因此次事件支付的前期医疗费暂计{medical_amount:.2f}元"
    _write_item(seq, "医疗费", medical_basis, medical_amount)
    seq += 1

    # 2. 护理费（可能有多行计算依据）
    nursing_start = row_idx
    if nursing_amount > 0 or nursing_supplies > 0:
        nursing_basis_1 = ""
        if nursing_supplies > 0:
            nursing_basis_1 = f"原告方因此次事件支付的医护用品费用{nursing_supplies:.2f}元"
        nursing_total = nursing_amount + nursing_supplies
        _write_item(seq, "护理费", nursing_basis_1, nursing_total)

        # 第二行：住院护理费计算公式
        if stay_days > 0:
            formula = (
                f"原告{analysis_result.get('patient_name', '')}共住院治疗{stay_days}天"
                f"（{admission_date}至{discharge_date}），其护理费按云南省上一年度"
                f"居民服务、修理和其他服务业非私营单位在岗职工年平均工资52940元/年计算："
                f"52940元/年÷365天×{stay_days}天"
            )
            _write_continuation(formula)
            _merge_seq_project(nursing_start, row_idx - 1)
    else:
        _write_item(seq, "护理费", "", 0)
    seq += 1

    # 3. 住院伙食补助费
    food_basis = ""
    if stay_days > 0:
        food_basis = (
            f"原告{analysis_result.get('patient_name', '')}共住院治疗{stay_days}天"
            f"（{admission_date}至{discharge_date}），按云南省标准100元/天计算："
            f"100元/天×{stay_days}天"
        )
    _write_item(seq, "住院伙食补助费", food_basis, food_amount)
    seq += 1

    # 4. 营养费
    nutrition_basis = ""
    if stay_days > 0:
        nutrition_basis = (
            f"原告{analysis_result.get('patient_name', '')}共住院治疗{stay_days}天"
            f"（{admission_date}至{discharge_date}），按50元/天计算："
            f"50元/天×{stay_days}天"
        )
    _write_item(seq, "营养费", nutrition_basis, nutrition_amount)
    seq += 1

    # 5. 交通住宿费
    _write_item(seq, "交通住宿费", "", transport_amount)
    seq += 1

    # 6. 鉴定费
    appraisal_basis = "尚未鉴定，暂不计算" if appraisal_amount == 0 else "凭票据实报实销"
    _write_item(seq, "鉴定费", appraisal_basis, appraisal_amount)
    seq += 1

    # 7. 精神损害抚慰金
    _write_item(seq, "精神损害抚慰金", "", spiritual_amount)
    seq += 1

    # ── 总计行 ──
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
    ws.cell(row=row_idx, column=1, value="总计")
    ws.cell(row=row_idx, column=1).font = _FANGSONG_BOLD
    ws.cell(row=row_idx, column=1).alignment = _CENTER_ALIGN
    ws.cell(row=row_idx, column=4, value=total_amount)
    ws.cell(row=row_idx, column=4).font = _FANGSONG_BOLD
    ws.cell(row=row_idx, column=4).number_format = _MONEY_FORMAT
    ws.cell(row=row_idx, column=4).alignment = _CENTER_ALIGN
    for col in range(1, COL_COUNT + 1):
        ws.cell(row=row_idx, column=col).border = _THIN_BORDER
        ws.cell(row=row_idx, column=col).fill = _TOTAL_FILL
    row_idx += 1

    # ── 备注行 ──
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=COL_COUNT)
    note_cell = ws.cell(row=row_idx, column=1,
                        value="备注：因本案尚未鉴定，故上述各项赔偿费用及残疾赔偿金等费用待鉴定后再行补充变更。")
    note_cell.font = _FANGSONG_CELL
    note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for col in range(1, COL_COUNT + 1):
        ws.cell(row=row_idx, column=col).border = _THIN_BORDER
    row_idx += 1

    # ── 列宽 ──
    col_widths = [8, 18, 60, 16]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1

    # ── 追加各项费用明细表（每个有票据数据的费用项一个 sheet）──
    # 失败不影响主清单生成。
    try:
        _append_fee_detail_sheets(wb, catalog_data)
    except Exception as e:
        logger.warning(f"Failed to append fee detail sheets: {type(e).__name__}: {e}")

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_fee_details_inline_data(catalog_data: dict, analysis_result: dict) -> dict[str, bytes]:
    """医疗费用汇总表 Excel（按医院分组，逐条明细）

    从 catalog_data.groups 中提取 fee_receipt 分类的材料，
    按医院名称分组，展示每张票据的日期、金额、报销/自费等信息。

    返回 {"医疗费用汇总表": bytes}
    """
    results: dict[str, bytes] = {}

    # ── 收集所有 fee_receipt 材料 ──
    fee_items = _collect_fee_receipt_items(catalog_data)

    if not fee_items:
        # 没有费用票据数据时，生成空表
        wb = Workbook()
        ws = wb.active
        ws.title = "医疗费用汇总表"
        ws.merge_cells("A1:J1")
        ws.cell(row=1, column=1, value="医疗费用汇总表（无费用票据数据）").font = _FANGSONG_TITLE
        output = io.BytesIO()
        wb.save(output)
        results["医疗费用汇总表"] = output.getvalue()
        return results

    # ── 按医院名称分组 ──
    hospital_groups: dict[str, list[dict]] = {}
    for item in fee_items:
        hospital = item.get("hospital_name", "") or "未知医院"
        if hospital not in hospital_groups:
            hospital_groups[hospital] = []
        hospital_groups[hospital].append(item)

    wb = Workbook()
    ws = wb.active
    ws.title = "医疗费用汇总表"

    case_name = analysis_result.get("case_name", "案件")

    # ── 标题行 ──
    ws.merge_cells("A1:J1")
    title_cell = ws.cell(row=1, column=1, value=f"{case_name} — 医疗费用汇总表")
    title_cell.font = _FANGSONG_TITLE
    title_cell.alignment = _CENTER_ALIGN

    # ── 表头（行3）──
    HEADER_ROW = 3
    headers = [
        "序号", "医院名称", "名称/项目", "日期", "金额（元）",
        "报销金额", "个人自费", "类型", "住院天数", "票据尾号",
    ]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=HEADER_ROW, column=col_idx, value=header)
    _apply_header_style(ws, HEADER_ROW, len(headers))

    # ── 按医院分组输出 ──
    row_idx = HEADER_ROW + 1
    seq = 1
    grand_total = 0.0
    grand_insurance = 0.0
    grand_out_of_pocket = 0.0

    for hospital_name, items in hospital_groups.items():
        # 医院名子标题行
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
        hospital_cell = ws.cell(row=row_idx, column=1, value=f"▶ {hospital_name}")
        hospital_cell.font = _BOLD_FONT
        hospital_cell.fill = _SUBTITLE_FILL
        hospital_cell.alignment = _LEFT_ALIGN
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = _THIN_BORDER
            ws.cell(row=row_idx, column=col).fill = _SUBTITLE_FILL
        row_idx += 1

        # 该医院下的逐条票据
        sub_total = 0.0
        sub_insurance = 0.0
        sub_out_of_pocket = 0.0

        for item in items:
            amount = item.get("total_amount", 0) or 0
            insurance = item.get("insurance_amount", 0) or 0
            out_of_pocket = item.get("out_of_pocket", 0) or 0
            date_str = item.get("date", "")
            title = item.get("title", "")
            receipt_type = item.get("receipt_type", "门诊")
            stay_days = item.get("stay_days", "")
            receipt_tail = item.get("receipt_tail", "")

            sub_total += amount
            sub_insurance += insurance
            sub_out_of_pocket += out_of_pocket

            ws.cell(row=row_idx, column=1, value=seq)
            ws.cell(row=row_idx, column=2, value="")  # 医院名已在子标题行
            ws.cell(row=row_idx, column=3, value=title)
            ws.cell(row=row_idx, column=4, value=date_str)
            ws.cell(row=row_idx, column=5, value=amount)
            ws.cell(row=row_idx, column=6, value=insurance if insurance else "")
            ws.cell(row=row_idx, column=7, value=out_of_pocket if out_of_pocket else "")
            ws.cell(row=row_idx, column=8, value=receipt_type)
            ws.cell(row=row_idx, column=9, value=stay_days)
            ws.cell(row=row_idx, column=10, value=receipt_tail)
            _apply_cell_style(ws, row_idx, len(headers), money_cols={5, 6, 7})
            row_idx += 1
            seq += 1

        # 小计行
        grand_total += sub_total
        grand_insurance += sub_insurance
        grand_out_of_pocket += sub_out_of_pocket

        ws.cell(row=row_idx, column=3, value=f"{hospital_name} 小计")
        ws.cell(row=row_idx, column=3).font = _BOLD_FONT
        ws.cell(row=row_idx, column=5, value=sub_total)
        ws.cell(row=row_idx, column=5).font = _BOLD_FONT
        ws.cell(row=row_idx, column=5).number_format = _MONEY_FORMAT
        ws.cell(row=row_idx, column=6, value=sub_insurance if sub_insurance else "")
        ws.cell(row=row_idx, column=6).font = _BOLD_FONT
        ws.cell(row=row_idx, column=6).number_format = _MONEY_FORMAT
        ws.cell(row=row_idx, column=7, value=sub_out_of_pocket if sub_out_of_pocket else "")
        ws.cell(row=row_idx, column=7).font = _BOLD_FONT
        ws.cell(row=row_idx, column=7).number_format = _MONEY_FORMAT
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = _THIN_BORDER
            ws.cell(row=row_idx, column=col).fill = _SUBTOTAL_FILL
        row_idx += 1

    # ── 总合计行 ──
    ws.cell(row=row_idx, column=3, value="总  计")
    ws.cell(row=row_idx, column=3).font = _FANGSONG_BOLD
    ws.cell(row=row_idx, column=5, value=grand_total)
    ws.cell(row=row_idx, column=5).font = _FANGSONG_BOLD
    ws.cell(row=row_idx, column=5).number_format = _MONEY_FORMAT
    ws.cell(row=row_idx, column=6, value=grand_insurance if grand_insurance else "")
    ws.cell(row=row_idx, column=6).font = _FANGSONG_BOLD
    ws.cell(row=row_idx, column=6).number_format = _MONEY_FORMAT
    ws.cell(row=row_idx, column=7, value=grand_out_of_pocket if grand_out_of_pocket else "")
    ws.cell(row=row_idx, column=7).font = _FANGSONG_BOLD
    ws.cell(row=row_idx, column=7).number_format = _MONEY_FORMAT
    for col in range(1, len(headers) + 1):
        ws.cell(row=row_idx, column=col).border = _THIN_BORDER
        ws.cell(row=row_idx, column=col).fill = _TOTAL_FILL

    # ── 列宽 ──
    col_widths = [8, 22, 28, 18, 16, 14, 14, 10, 10, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ── 打印设置 ──
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1

    output = io.BytesIO()
    wb.save(output)
    results["医疗费用汇总表"] = output.getvalue()

    return results


def _collect_fee_receipt_items(catalog_data: dict) -> list[dict]:
    """从 catalog_data.groups 中提取 fee_receipt 分类的材料明细

    返回列表，每项包含：
      hospital_name, title, date, total_amount,
      insurance_amount, out_of_pocket,
      receipt_type(住院/门诊), stay_days, receipt_tail
    """
    items: list[dict] = []

    for group in catalog_data.get("groups", []):
        category = group.get("category", "")
        if category != "fee_receipt":
            continue

        for mat_item in group.get("items", []):
            fees = mat_item.get("fees", {}) or {}
            identity = mat_item.get("identity", {}) or {}
            evidence_name = mat_item.get("evidence_name", {}) or {}
            raw_extracted = mat_item.get("raw_extracted", {}) or {}
            layer3 = raw_extracted.get("layer_3_treatment", {}) or {}

            # 总额
            total_amount = fees.get("total_amount", 0) or 0

            # 如果 total_amount 为 0，尝试从 items 汇总
            if total_amount <= 0:
                fee_items_list = fees.get("items", []) or []
                for fi in fee_items_list:
                    a = fi.get("amount", 0) or 0
                    # 尝试将字符串转换为数字
                    if isinstance(a, str):
                        try:
                            a = float(a)
                        except (ValueError, TypeError):
                            a = 0
                    if isinstance(a, (int, float)) and a > 0:
                        total_amount += a

            # 医院名
            hospital_name = identity.get("hospital_name", "")

            # 日期
            date_str = evidence_name.get("date", "")
            # 尝试从诊疗经过获取入院-出院日期
            admission = layer3.get("admission_date", "")
            discharge = layer3.get("discharge_date", "")

            # 判断住院/门诊
            stay_days = ""
            receipt_type = "门诊"
            if admission and discharge:
                receipt_type = "住院"
                try:
                    from datetime import datetime
                    d1 = datetime.strptime(admission[:10], "%Y-%m-%d")
                    d2 = datetime.strptime(discharge[:10], "%Y-%m-%d")
                    stay_days = (d2 - d1).days
                    if stay_days < 0:
                        stay_days = ""
                except (ValueError, IndexError):
                    stay_days = ""
                date_str = f"{admission[:10]} ~ {discharge[:10]}"

            # 报销/自费
            insurance_amount = fees.get("insurance_amount") or 0
            if isinstance(insurance_amount, str):
                try:
                    insurance_amount = float(insurance_amount)
                except (ValueError, TypeError):
                    insurance_amount = 0
            out_of_pocket = fees.get("out_of_pocket") or 0
            if isinstance(out_of_pocket, str):
                try:
                    out_of_pocket = float(out_of_pocket)
                except (ValueError, TypeError):
                    out_of_pocket = 0

            # 票据尾号 — 从原始文件名取后4位数字
            original_filename = _clean_text(evidence_name.get("original_filename", "")) or mat_item.get("title", "")
            receipt_tail = _extract_receipt_tail(original_filename)

            # 标题 — 清理JSON污染
            title = _clean_text(evidence_name.get("title", "")) or _clean_text(mat_item.get("title", ""))

            # 医院名 — 清理JSON污染
            hospital_name = _clean_text(hospital_name)

            # 日期 — 清理JSON污染
            date_str = _clean_text(date_str)

            items.append({
                "hospital_name": hospital_name,
                "title": title,
                "date": date_str,
                "total_amount": total_amount,
                "insurance_amount": insurance_amount,
                "out_of_pocket": out_of_pocket,
                "receipt_type": receipt_type,
                "stay_days": stay_days,
                "receipt_tail": receipt_tail,
            })

    return items


def _extract_receipt_tail(filename: str) -> str:
    """从文件名中提取票据尾号（取最后4位数字）"""
    import re
    digits = re.findall(r"\d", filename)
    if len(digits) >= 4:
        return "".join(digits[-4:])
    return ""


def generate_compensation_calculation_excel(compensation_data: dict | None, case_name: str = "",
                                              plaintiff_name: str = "", case_type: str = "injury") -> bytes:
    """生成赔偿费用清单 Excel（标准法律文书格式，参照手工模板）

    格式与手工填写的赔偿费用清单一致：
    - 标题行：赔偿费用清单
    - 表头：序号 | 项目 | 计算依据 | 金额（元）
    - 各项赔偿费用（金额为 0 的项目跳过，不输出空行）
    - 护理费子行：上方金额行含用品发票小计，下方子行写住院天数计算过程
    - 合计行：用 SUM 公式（对金额列所有数据行求和）
    - 备注行

    当 compensation_data 为空或没有 items 时，按案件类型生成模板：
    项目名称和计算依据照出，金额列留白（显示横线）。

    Args:
        compensation_data: calculate_all() 返回的赔偿计算结果字典，可为 None/空
        case_name: 案件名称
        plaintiff_name: 原告姓名
        case_type: 案件类型 injury/death
    """
    compensation_data = compensation_data or {}
    items = compensation_data.get('items', [])
    params = compensation_data.get('params', {})
    has_data = bool(items)

    wb = Workbook()
    ws = wb.active
    ws.title = "赔偿费用清单"

    # 样式
    title_font = Font(name='宋体', size=14, bold=True)
    header_font = Font(name='宋体', size=11, bold=True)
    cell_font = Font(name='宋体', size=11)
    money_fmt = '#,##0.00'
    bold_font = Font(name='宋体', size=11, bold=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # 列宽（4列：序号、项目、计算依据、金额）
    col_widths = [8, 20, 55, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── 标题行 ──
    ws.merge_cells('A1:D1')
    ws['A1'] = '赔偿费用清单'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    # ── 表头（行2）──
    headers = ['序号', '项目', '计算依据', '金额（元）']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    row_idx = 3
    seq = 1
    # 记录金额列的数据行号，用于合计 SUM 公式
    amount_rows: list[int] = []

    if has_data:
        # ── 有计算数据：只输出金额 > 0 或 manual_amount 不为 null 的项目 ──
        for item in items:
            fee_type = item.get('fee_type', '')
            fee_name = item.get('fee_name', '')
            amount = item.get('manual_amount') or item.get('amount', 0)
            basis = item.get('calculation_basis', '')
            try:
                amount = float(str(amount))
            except (ValueError, TypeError):
                amount = 0.0

            # dependent_living 不在清单中单独列出（已包含在残疾/死亡赔偿金中）
            if fee_type == 'dependent_living':
                continue

            ws.cell(row=row_idx, column=1, value=seq).font = cell_font
            ws.cell(row=row_idx, column=1).border = thin_border
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center')

            ws.cell(row=row_idx, column=2, value=fee_name).font = cell_font
            ws.cell(row=row_idx, column=2).border = thin_border

            # 护理费主行：basis 写护理用品发票部分，金额用完整金额
            # 非护理费正常写 basis
            if fee_type == 'nursing_fee':
                supplies_basis = _get_nursing_supplies_basis(item)
                ws.cell(row=row_idx, column=3, value=supplies_basis).font = cell_font
            else:
                ws.cell(row=row_idx, column=3, value=basis).font = cell_font
            ws.cell(row=row_idx, column=3).border = thin_border

            cell_amount = ws.cell(row=row_idx, column=4, value=amount)
            cell_amount.number_format = money_fmt
            cell_amount.font = cell_font
            cell_amount.border = thin_border
            cell_amount.alignment = Alignment(horizontal='center')
            amount_rows.append(row_idx)

            row_idx += 1

            # 护理费子行：写住院天数计算过程，序号/项目/金额列合并
            if fee_type == 'nursing_fee':
                sub_detail = _get_nursing_sub_detail(params, plaintiff_name)
                if sub_detail:
                    # 合并序号、项目、金额列
                    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=1)
                    ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=2)
                    ws.merge_cells(start_row=row_idx, start_column=4, end_row=row_idx, end_column=4)
                    ws.cell(row=row_idx, column=3, value=sub_detail).font = cell_font
                    for c in range(1, 5):
                        ws.cell(row=row_idx, column=c).border = thin_border
                    row_idx += 1
            else:
                # 非护理费子行详细计算依据
                sub_row = _get_calculation_detail(fee_type, params, plaintiff_name)
                if sub_row:
                    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=1)
                    ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=2)
                    ws.merge_cells(start_row=row_idx, start_column=4, end_row=row_idx, end_column=4)
                    ws.cell(row=row_idx, column=3, value=sub_row).font = cell_font
                    for c in range(1, 5):
                        ws.cell(row=row_idx, column=c).border = thin_border
                    row_idx += 1

            seq += 1

        # 合计行：用 SUM 公式
        if amount_rows:
            sum_parts = "+".join(f"D{r}" for r in amount_rows)
            sum_formula = f"={sum_parts}"
        else:
            sum_formula = 0
    else:
        # ── 无计算数据：按案件类型生成模板，金额留白 ──
        template_items = _get_template_items(case_type)
        for fee_name, basis in template_items:
            ws.cell(row=row_idx, column=1, value=seq).font = cell_font
            ws.cell(row=row_idx, column=1).border = thin_border
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center')

            ws.cell(row=row_idx, column=2, value=fee_name).font = cell_font
            ws.cell(row=row_idx, column=2).border = thin_border

            ws.cell(row=row_idx, column=3, value=basis).font = cell_font
            ws.cell(row=row_idx, column=3).border = thin_border

            # 金额留白（横线占位）
            cell_amount = ws.cell(row=row_idx, column=4, value="——")
            cell_amount.font = cell_font
            cell_amount.border = thin_border
            cell_amount.alignment = Alignment(horizontal='center')

            row_idx += 1
            seq += 1

        sum_formula = "——"

    # ── 合计行 ──
    ws.cell(row=row_idx, column=2, value='总计').font = bold_font
    ws.cell(row=row_idx, column=2).border = thin_border
    ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=row_idx, column=1).border = thin_border
    ws.cell(row=row_idx, column=3).border = thin_border
    cell_total = ws.cell(row=row_idx, column=4, value=sum_formula)
    if isinstance(sum_formula, str) and sum_formula.startswith('='):
        cell_total.number_format = money_fmt
    cell_total.font = bold_font
    cell_total.border = thin_border
    cell_total.alignment = Alignment(horizontal='center')
    row_idx += 1

    # ── 备注行 ──
    if case_type == "injury":
        note = "备注：因本案尚未鉴定，故上述各项赔偿费用及残疾赔偿金等费用待鉴定后再行补充变更。"
    else:
        note = "备注：因本案尚未鉴定，故上述各项赔偿费用及鉴定费等，待鉴定后再行补充变更。"
    ws.merge_cells(f'A{row_idx}:D{row_idx}')
    ws.cell(row=row_idx, column=1, value=note).font = cell_font

    # ── 打印设置 ──
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _get_nursing_supplies_basis(item: dict) -> str:
    """生成护理费主行的计算依据（护理用品发票部分）"""
    sources = item.get('sources', [])
    if sources:
        filenames = [s.get('filename', '') for s in sources if s.get('filename')]
        if filenames:
            return f"原告方因此次事件支付的医护用品费用合计（见{'、'.join(filenames[:3])}）"
    return "原告方因此次事件支付的医护用品费用"


def _get_nursing_sub_detail(params: dict, plaintiff_name: str = "") -> str:
    """生成护理费子行的详细计算依据（住院天数 × 日护理费）"""
    try:
        hospital_days = int(float(params.get('nursing_days', 0) or params.get('hospital_days', 0)))
        annual_salary = float(params.get('nursing_annual_salary', 0))
        if annual_salary == 0:
            legacy = params.get('nursing_monthly_salary') or params.get('monthly_salary') or 8500
            annual_salary = float(legacy) * 12
        person_count = int(params.get('nursing_person_count', 1) or 1)
        dep_level = str(params.get('nursing_dependency_level') or 'full')
    except (ValueError, TypeError):
        return ""

    if hospital_days <= 0:
        return ""

    name = plaintiff_name or "原告"
    # 使用全精度日费率用于展示（与计算引擎一致）
    daily = annual_salary / 365

    dep_map = {"full": "完全护理依赖(100%)", "mostly": "大部分护理依赖(80%)", "partial": "部分护理依赖(50%)"}
    dep_label = dep_map.get(dep_level, "")

    person_part = f"× {person_count}人" if person_count > 1 else ""
    dep_part = f"× {dep_label}" if dep_level != "full" else ""

    return f"{name}共住院治疗{hospital_days}天，其护理费按上一年度居民服务、修理和其他服务业非私营单位在岗职工年平均工资{annual_salary:.0f}元/年计算：{annual_salary:.0f}元/年÷365天×{hospital_days}天{person_part}{dep_part}"


def _get_template_items(case_type: str) -> list[tuple[str, str]]:
    """按案件类型返回模板赔偿项目（名称 + 计算依据），用于无数据时的空模板生成"""
    if case_type == "death":
        return [
            ("医疗费", "凭票据实报实销"),
            ("误工费", "误工收入 × 误工天数（参照鉴定意见）"),
            ("护理费", "住院期间护理费 = 护理标准 × 住院天数"),
            ("住院伙食补助费", "住院伙食补助标准 × 住院天数"),
            ("营养费", "营养费标准 × 营养期天数（参照鉴定意见）"),
            ("死亡赔偿金", "受诉法院所在地上年度城镇居民人均可支配收入 × 20年（含被扶养人生活费）"),
            ("丧葬费", "受诉法院所在地上年度职工月平均工资 × 6个月"),
            ("交通费", "凭票据实报实销"),
            ("精神损害抚慰金", "根据死亡后果酌定"),
        ]
    else:
        # injury（含新生儿）
        return [
            ("医疗费", "凭票据实报实销"),
            ("误工费", "误工收入 × 误工天数（参照鉴定意见）"),
            ("护理费", "住院期间护理费 = 护理标准 × 住院天数"),
            ("住院伙食补助费", "住院伙食补助标准 × 住院天数"),
            ("营养费", "营养费标准 × 营养期天数（参照鉴定意见）"),
            ("残疾赔偿金", "受诉法院所在地上年度城镇居民人均可支配收入 × 20年 × 伤残系数（含被扶养人生活费）"),
            ("后续治疗费", "参照鉴定意见或后续实际发生费用"),
            ("交通费", "凭票据实报实销"),
            ("精神损害抚慰金", "根据伤残等级酌定"),
        ]


def _get_calculation_detail(fee_type: str, params: dict, plaintiff_name: str = "") -> str:
    """生成子行的详细计算依据说明（参照手工模板格式）

    仅对部分需要详细说明的项目生成子行，其他返回空字符串。
    护理费由 _get_nursing_sub_detail 独立处理，此函数不再处理。
    """
    try:
        hospital_days = int(float(params.get('hospital_days', 0)))
        annual_income = float(params.get('annual_income', 49283))
        monthly_salary = float(params.get('monthly_salary', 8500))
        daily_food = float(params.get('daily_food_subsidy', 100))
        daily_nutrition = float(params.get('daily_nutrition', 50))
        compensation_years = int(float(params.get('compensation_years', 20)))
        disability_coeff = float(params.get('disability_coefficient', 1.0))
        lost_wage_days = int(float(params.get('lost_wage_days', 0)))
        victim_age = int(float(params.get('victim_age', 0)))
    except (ValueError, TypeError):
        return ""

    name = plaintiff_name or "原告"

    if fee_type == 'lost_wages' and lost_wage_days > 0:
        daily = monthly_salary / 30
        return f"{name}的误工费按上一年度职工月平均工资{monthly_salary:.0f}元/月计算：{monthly_salary:.0f}元/月÷30天×{lost_wage_days}天"

    if fee_type == 'food_subsidy' and hospital_days > 0:
        return f"{name}共住院治疗{hospital_days}天，按{daily_food:.0f}元/天计算：{daily_food:.0f}元/天×{hospital_days}天"

    if fee_type == 'nutrition_fee' and hospital_days > 0:
        return f"{name}共住院治疗{hospital_days}天，按{daily_nutrition:.0f}元/天计算：{daily_nutrition:.0f}元/天×{hospital_days}天"

    if fee_type == 'disability_compensation' and compensation_years > 0:
        basis = f"按上一年度城镇居民人均可支配收入{annual_income:.0f}元/年×{compensation_years}年×伤残系数{disability_coeff}"
        # 含被扶养人生活费标注
        annual_consumption = float(params.get('annual_consumption', 0))
        if annual_consumption > 0:
            dep_amt = annual_consumption * compensation_years * disability_coeff
            basis += f"（含被扶养人生活费 ¥{dep_amt:.2f}）"
        return basis

    if fee_type == 'death_compensation':
        # 死亡赔偿金年龄递减
        if victim_age > 0:
            if victim_age >= 75:
                actual_years = 5
            elif victim_age >= 60:
                actual_years = max(20 - (victim_age - 60), 5)
            else:
                actual_years = compensation_years if compensation_years > 0 else 20
        else:
            actual_years = compensation_years
        if actual_years > 0:
            basis = f"按上一年度城镇居民人均可支配收入{annual_income:.0f}元/年×{actual_years}年"
            annual_consumption = float(params.get('annual_consumption', 0))
            if annual_consumption > 0:
                dep_amt = annual_consumption * actual_years
                basis += f"（含被扶养人生活费 ¥{dep_amt:.2f}）"
            return basis

    if fee_type == 'funeral_fee':
        return f"按上一年度职工月平均工资{monthly_salary:.0f}元/月，以六个月总额计算：{monthly_salary:.0f}元/月×6月"

    return ""


# ═══════════════════════════════════════════════════════════════════════════════
# 10项赔偿明细 ZIP 包生成
# ═══════════════════════════════════════════════════════════════════════════════

# 10项赔偿明细 ZIP 包中的文件名顺序与 fee_type 映射
ZIP_FEE_ITEMS_ORDER: list[tuple[str, str]] = [
    # (文件名中文名, 对应 compensation_data.items 中的 fee_type)
    ("医疗费", "medical_fee"),
    ("误工费", "lost_wages"),
    ("护理费", "nursing_fee"),
    ("住院伙食补助费", "food_subsidy"),
    ("营养费", "nutrition_fee"),
    ("残疾赔偿金", "disability_compensation"),
    ("死亡赔偿金", "death_compensation"),
    ("交通住宿费", "transport_fee"),
    ("鉴定费", "appraisal_fee"),
    ("精神损害抚慰金", "spiritual_damage"),
    ("被扶养人生活费", "dependent_living"),
]


def _collect_fee_receipt_items_v2(catalog_data: dict) -> list[dict]:
    """从 catalog_data.groups 中提取 fee_receipt 分类材料明细（支持新版带 fee_type 字段的结构）。

    返回列表，每项包含：
      fee_type, hospital_name, title, date, amount, insurance_amount,
      self_pay_amount, receipt_type, stay_days, receipt_tail, evidence_page
    """
    items: list[dict] = []

    for group in catalog_data.get("groups", []):
        if group.get("category") != "fee_receipt":
            continue

        for mat_item in group.get("items", []):
            # 兼容新旧两种数据结构
            fee_type = mat_item.get("fee_type", "") or ""
            fees = mat_item.get("fees", {}) or {}
            identity = mat_item.get("identity", {}) or {}
            evidence_name = mat_item.get("evidence_name", {}) or {}
            raw_extracted = mat_item.get("raw_extracted", {}) or {}
            layer3 = raw_extracted.get("layer_3_treatment", {}) or {}
            layer4 = raw_extracted.get("layer_4_fees", {}) or (fees or {})

            # 金额
            amount = mat_item.get("amount") or fees.get("total_amount", 0) or 0
            if not isinstance(amount, (int, float)):
                try:
                    amount = float(str(amount).replace(",", ""))
                except (ValueError, TypeError):
                    amount = 0.0

            # 如果 amount 为 0，尝试从 items 汇总
            if amount <= 0:
                fee_items_list = layer4.get("items", []) or fees.get("items", []) or []
                for fi in fee_items_list:
                    a = fi.get("amount", 0) or 0
                    if isinstance(a, str):
                        try:
                            a = float(a.replace(",", ""))
                        except (ValueError, TypeError):
                            a = 0
                    if isinstance(a, (int, float)) and a > 0:
                        amount += a

            # 医院名
            hospital_name = _clean_text(identity.get("hospital_name", ""))

            # 日期
            date_str = _clean_text(evidence_name.get("date", "") or mat_item.get("date", ""))
            admission = layer3.get("admission_date", "")
            discharge = layer3.get("discharge_date", "")

            # 住院/门诊
            stay_days: Any = ""
            receipt_type = _clean_text(mat_item.get("fee_subtype", "")) or "门诊"
            if admission and discharge:
                receipt_type = "住院"
                try:
                    from datetime import datetime as _dt
                    d1 = _dt.strptime(admission[:10], "%Y-%m-%d")
                    d2 = _dt.strptime(discharge[:10], "%Y-%m-%d")
                    stay_days = (d2 - d1).days
                    if stay_days < 0:
                        stay_days = ""
                    date_str = f"{admission[:10]} ~ {discharge[:10]}"
                except (ValueError, IndexError):
                    stay_days = ""

            # 报销/自费
            insurance_amount = mat_item.get("insurance_amount") or fees.get("insurance_amount") or 0
            if isinstance(insurance_amount, str):
                try:
                    insurance_amount = float(insurance_amount.replace(",", ""))
                except (ValueError, TypeError):
                    insurance_amount = 0.0
            self_pay_amount = mat_item.get("self_pay_amount") or fees.get("out_of_pocket") or 0
            if isinstance(self_pay_amount, str):
                try:
                    self_pay_amount = float(self_pay_amount.replace(",", ""))
                except (ValueError, TypeError):
                    self_pay_amount = 0.0

            # 票据尾号
            original_filename = _clean_text(evidence_name.get("original_filename", "")) or mat_item.get("title", "")
            receipt_tail = _extract_receipt_tail(original_filename)

            # 标题
            title = _clean_text(evidence_name.get("title", "")) or _clean_text(mat_item.get("title", ""))

            # 证据页码（如存在）
            evidence_page = mat_item.get("evidence_page", "") or ""

            items.append({
                "fee_type": fee_type,
                "hospital_name": hospital_name,
                "title": title,
                "date": date_str,
                "amount": float(amount),
                "insurance_amount": float(insurance_amount),
                "self_pay_amount": float(self_pay_amount),
                "receipt_type": receipt_type,
                "stay_days": stay_days,
                "receipt_tail": receipt_tail,
                "evidence_page": evidence_page,
                "ocr_text": mat_item.get("ocr_text", "") or "",
                "evidence_name": evidence_name,
                "material_id": mat_item.get("material_id", ""),
            })

    return items


def _to_float(value: Any, default: float = 0.0) -> float:
    """安全转换为 float，处理字符串/Decimal/None。"""
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", ""))
    except (ValueError, TypeError):
        return default


# ── 各项费用明细表（追加到赔偿费用清单工作簿）─────────────────────────────────
# fee_type → 显示名（来自 ZIP_FEE_ITEMS_ORDER）
_FEE_TYPE_DISPLAY = {ft: name for name, ft in ZIP_FEE_ITEMS_ORDER}


def _safe_sheet_title(base: str, used: set[str]) -> str:
    """生成合法且唯一的工作表名（Excel 限制：<=31 字符，不含 []:*?/\\）。"""
    name = re.sub(r"[\[\]:\*\?/\\]", "", base).strip() or "明细"
    name = name[:31]
    candidate = name
    i = 2
    while candidate in used:
        suffix = f"_{i}"
        candidate = name[: 31 - len(suffix)] + suffix
        i += 1
    used.add(candidate)
    return candidate


def _write_fee_detail_ws(ws: Any, sheet_title: str, items: list[dict]) -> None:
    """在给定工作表写入某一费用项的明细表（按医院分组 + 小计 + 总计）。

    items 为 _collect_fee_receipt_items_v2 返回的明细
    （含 amount / insurance_amount / self_pay_amount 等字段）。
    """
    headers = [
        "序号", "医院名称", "名称/项目", "日期", "金额（元）",
        "报销金额", "个人自费", "类型", "住院天数", "票据尾号",
    ]
    col_count = len(headers)

    # 标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    title_cell = ws.cell(row=1, column=1, value=sheet_title)
    title_cell.font = _FANGSONG_TITLE
    title_cell.alignment = _CENTER_ALIGN

    # 表头行
    header_row = 3
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=header_row, column=col_idx, value=header)
    _apply_header_style(ws, header_row, col_count)

    # 按医院分组
    hospital_groups: dict[str, list[dict]] = {}
    for it in items:
        hosp = it.get("hospital_name") or "未知医院"
        hospital_groups.setdefault(hosp, []).append(it)

    row_idx = header_row + 1
    seq = 1
    grand_total = grand_insurance = grand_self = 0.0

    for hosp, its in hospital_groups.items():
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=col_count)
        hc = ws.cell(row=row_idx, column=1, value=f"▶ {hosp}")
        hc.font = _BOLD_FONT
        hc.fill = _SUBTITLE_FILL
        hc.alignment = _LEFT_ALIGN
        for c in range(1, col_count + 1):
            ws.cell(row=row_idx, column=c).border = _THIN_BORDER
            ws.cell(row=row_idx, column=c).fill = _SUBTITLE_FILL
        row_idx += 1

        sub_total = sub_insurance = sub_self = 0.0
        for it in its:
            amount = _to_float(it.get("amount"))
            insurance = _to_float(it.get("insurance_amount"))
            self_pay = _to_float(it.get("self_pay_amount"))
            sub_total += amount
            sub_insurance += insurance
            sub_self += self_pay
            ws.cell(row=row_idx, column=1, value=seq)
            ws.cell(row=row_idx, column=2, value="")
            ws.cell(row=row_idx, column=3, value=it.get("title", ""))
            ws.cell(row=row_idx, column=4, value=it.get("date", ""))
            ws.cell(row=row_idx, column=5, value=amount if amount else "")
            ws.cell(row=row_idx, column=6, value=insurance if insurance else "")
            ws.cell(row=row_idx, column=7, value=self_pay if self_pay else "")
            ws.cell(row=row_idx, column=8, value=it.get("receipt_type", ""))
            ws.cell(row=row_idx, column=9, value=it.get("stay_days", ""))
            ws.cell(row=row_idx, column=10, value=it.get("receipt_tail", ""))
            _apply_cell_style(ws, row_idx, col_count, money_cols={5, 6, 7})
            row_idx += 1
            seq += 1

        grand_total += sub_total
        grand_insurance += sub_insurance
        grand_self += sub_self

        ws.cell(row=row_idx, column=3, value=f"{hosp} 小计")
        ws.cell(row=row_idx, column=3).font = _BOLD_FONT
        ws.cell(row=row_idx, column=5, value=sub_total)
        ws.cell(row=row_idx, column=5).font = _BOLD_FONT
        ws.cell(row=row_idx, column=5).number_format = _MONEY_FORMAT
        if sub_insurance:
            ws.cell(row=row_idx, column=6, value=sub_insurance)
            ws.cell(row=row_idx, column=6).font = _BOLD_FONT
            ws.cell(row=row_idx, column=6).number_format = _MONEY_FORMAT
        if sub_self:
            ws.cell(row=row_idx, column=7, value=sub_self)
            ws.cell(row=row_idx, column=7).font = _BOLD_FONT
            ws.cell(row=row_idx, column=7).number_format = _MONEY_FORMAT
        for c in range(1, col_count + 1):
            ws.cell(row=row_idx, column=c).border = _THIN_BORDER
            ws.cell(row=row_idx, column=c).fill = _SUBTOTAL_FILL
        row_idx += 1

    # 总计行
    ws.cell(row=row_idx, column=3, value="总  计")
    ws.cell(row=row_idx, column=3).font = _FANGSONG_BOLD
    ws.cell(row=row_idx, column=5, value=grand_total)
    ws.cell(row=row_idx, column=5).font = _FANGSONG_BOLD
    ws.cell(row=row_idx, column=5).number_format = _MONEY_FORMAT
    if grand_insurance:
        ws.cell(row=row_idx, column=6, value=grand_insurance)
        ws.cell(row=row_idx, column=6).font = _FANGSONG_BOLD
        ws.cell(row=row_idx, column=6).number_format = _MONEY_FORMAT
    if grand_self:
        ws.cell(row=row_idx, column=7, value=grand_self)
        ws.cell(row=row_idx, column=7).font = _FANGSONG_BOLD
        ws.cell(row=row_idx, column=7).number_format = _MONEY_FORMAT
    for c in range(1, col_count + 1):
        ws.cell(row=row_idx, column=c).border = _THIN_BORDER
        ws.cell(row=row_idx, column=c).fill = _TOTAL_FILL

    col_widths = [8, 22, 28, 18, 16, 14, 14, 10, 10, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1


def _append_fee_detail_sheets(wb: Workbook, catalog_data: dict) -> None:
    """为赔偿费用清单工作簿追加各项费用明细表。

    每个有票据明细数据的费用项（按 fee_type 分组）追加一个明细工作表；
    无 fee_type 标注的票据归入"费用明细"。
    """
    items = _collect_fee_receipt_items_v2(catalog_data)
    if not items:
        return

    by_type: dict[str, list[dict]] = {}
    for it in items:
        ft = it.get("fee_type") or ""
        by_type.setdefault(ft, []).append(it)

    # 输出顺序：先按标准赔偿项顺序，再其余 fee_type，最后无 fee_type
    ordered_types: list[str] = []
    for _name, ft in ZIP_FEE_ITEMS_ORDER:
        if ft in by_type and ft not in ordered_types:
            ordered_types.append(ft)
    for ft in by_type:
        if ft and ft not in ordered_types:
            ordered_types.append(ft)
    if "" in by_type:
        ordered_types.append("")

    used_titles: set[str] = {ws.title for ws in wb.worksheets}
    for ft in ordered_types:
        group_items = by_type.get(ft) or []
        if not group_items:
            continue
        display = _FEE_TYPE_DISPLAY.get(ft, "") if ft else ""
        full_title = f"{display}明细" if display else "费用明细"
        sheet_title = _safe_sheet_title(full_title, used_titles)
        ws = wb.create_sheet(title=sheet_title)
        _write_fee_detail_ws(ws, full_title, group_items)


def _write_title_row(ws: Any, title: str, col_count: int) -> int:
    """写入标题行（合并第一行），仿宋16号加粗居中，返回下一可用行号。"""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = _FANGSONG_TITLE
    cell.alignment = _CENTER_ALIGN
    return 2


def _write_header_row(ws: Any, headers: list[str], row: int) -> None:
    """写入表头行 — 仿宋12号加粗，居中对齐。"""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = _FANGSONG_HEADER
        cell.alignment = _CENTER_ALIGN
        cell.border = _THIN_BORDER


# ── 各类 Excel 生成函数 ──────────────────────────────────────────────────────

# （仿宋字体常量已统一移至文件头部样式定义区）


def _parse_medical_fee_from_ocr(ocr_text: str) -> dict:
    """从 OCR 文本中智能解析医疗费用信息。

    解析字段：
    - hospital_name: 医院名称（从"龙陵县人民医院"等关键词提取）
    - date: 日期（交易时间 / 开票日期）
    - amount: 金额（合计金额）
    - insurance_amount: 报销金额（统筹支付）
    - self_pay_amount: 个人自费（个人现金/扫码付）
    - receipt_type: 类型（门诊费/住院费/检查费等）
    - stay_days: 住院天数（从医保结算单提取）
    - receipt_tail: 票据尾号（票据号码后4位）

    Returns:
        dict with parsed fields
    """
    import re

    result = {
        "hospital_name": "",
        "date": "",
        "amount": 0.0,
        "insurance_amount": 0.0,
        "self_pay_amount": 0.0,
        "receipt_type": "",
        "stay_days": "",
        "receipt_tail": "",
    }

    if not ocr_text:
        return result

    text = ocr_text

    # ── 医院名称 ──
    hospital_patterns = [
        r'([\u4e00-\u9fa5]{2,8}县[\u4e00-\u9fa5]{2,6}(?:人民医院|医院|卫生院))',
        r'([\u4e00-\u9fa5]{2,8}市[\u4e00-\u9fa5]{2,6}(?:人民医院|医院))',
        r'([\u4e00-\u9fa5]{2,8}区[\u4e00-\u9fa5]{2,6}(?:人民医院|医院))',
        r'([\u4e00-\u9fa5]{2,10}(?:人民医院|中心医院|附属医院))',
    ]
    # 也从标题行提取（第一行通常是医院名称）
    first_line = text.strip().split('\n')[0].strip() if text.strip() else ""
    for pat in hospital_patterns:
        m = re.search(pat, text)
        if m:
            result["hospital_name"] = m.group(1)
            break
    # 如果第一行本身就是医院名（如"龙陵县人民医院住院一日清单"）
    if not result["hospital_name"] and first_line:
        for pat in hospital_patterns:
            m = re.search(pat, first_line)
            if m:
                result["hospital_name"] = m.group(1)
                break

    # ── 日期 ──
    # 交易时间：2025/8/18 14：56：14  或  交易时间：2025/9/2 17：55：30
    m = re.search(r'交易时间[：:]\s*(\d{4}[/\-]\d{1,2}[/\-]\d{1,2})', text)
    if m:
        result["date"] = m.group(1).replace('-', '/')
    else:
        # 费用发生日期：2025-12-16
        m = re.search(r'费用发生日期[：:]\s*\n?\s*(\d{4}[/\-]\d{1,2}[/\-]\d{1,2})', text)
        if m:
            result["date"] = m.group(1).replace('-', '/')
        else:
            # 开票日期：2026年03月24日
            m = re.search(r'开票日期[：:]\s*(\d{4})年(\d{1,2})月(\d{1,2})日', text)
            if m:
                result["date"] = f"{m.group(1)}/{int(m.group(2))}/{int(m.group(3))}"
            else:
                # 入院日期-出院日期（住院类）
                m_in = re.search(r'入院日期[：:]?\s*(\d{4}[/\-.]\d{1,2}[/\-.]\d{1,2})', text)
                m_out = re.search(r'出院日期[：:]?\s*(\d{4}[/\-.]\d{1,2}[/\-.]\d{1,2})', text)
                if m_in and m_out:
                    result["date"] = f"{m_in.group(1).replace('-','/')}-{m_out.group(1).replace('-','/')}"

    def _match_amount_after_label(text: str, labels: list[str]) -> float:
        """在text中查找任一label，匹配其前后数值（支持同行冒号、值跨行在上、值跨行在下）。

        医保结算单OCR中字段名与数值常跨行分布，且值可能在label上方或下方。
        例如:
            "257611.02\\n医疗费总额"      → 值在label上方
            "基金支付总额\\n217344.01"     → 值在label下方
            "合计金额：140元"              → 同行冒号
        """
        for label in labels:
            esc = re.escape(label)
            # 1) 同行：label：值 / label:值
            m = re.search(esc + r'[：:]\s*([\d,]+\.\d+|[\d,]+)', text)
            if m:
                return _to_float(m.group(1))
            # 2) 跨行值在下：label\n值
            m = re.search(esc + r'\s*\n\s*([\d,]+\.\d+|[\d,]+)', text)
            if m:
                return _to_float(m.group(1))
            # 3) 跨行值在上：值\nlabel
            m = re.search(r'([\d,]+\.\d+|[\d,]+)\s*\n\s*' + esc, text)
            if m:
                return _to_float(m.group(1))
        return 0.0

    # ── 金额（合计金额 / 医疗费总额）──
    # 1) 门诊小票：合计金额：140元  或  合计：271195.00元
    amt = _match_amount_after_label(text, ["合计金额", "合计"])
    if amt > 0:
        result["amount"] = amt
    # 2) 医保结算单：医疗费总额（可能跨行；优先取这个）
    amt_total = _match_amount_after_label(text, ["医疗费总额", "医疗费总预"])
    if amt_total > 0:
        # 结算单的总额始终优先（小票的"合计"可能会被误匹配为其它字段的合计）
        # 但只有当 OCR 文本含"结算单"或"医保"标识时才覆盖
        if "结算单" in text[:200] or "医疗保障" in text[:200]:
            result["amount"] = amt_total
        elif result["amount"] == 0:
            result["amount"] = amt_total
    # 3) 全自费金额（极少用，作为最后fallback，仅当上面都失败）
    if result["amount"] == 0:
        amt_self = _match_amount_after_label(text, ["全自费金额", "全自费金别"])
        if amt_self > 0:
            result["amount"] = amt_self

    # ── 报销金额（统筹支付 / 基金支付总额） ──
    ins = _match_amount_after_label(text, [
        "统筹支付",        # 门诊小票
        "基金支付总额",    # 医保结算单（主字段）
        "医保支付总额",    # OCR变体
    ])
    if ins > 0:
        result["insurance_amount"] = ins

    # ── 个人自费 ──
    # 1) 门诊小票字段
    self_pay = _match_amount_after_label(text, [
        "扫码付", "现金",
    ])
    # 2) 医保结算单字段："个人负担金额"包含现金+账户共济等，对齐原版口径(40267.01)
    if self_pay == 0:
        self_pay = _match_amount_after_label(text, [
            "个人负担金额",      # 医保结算单主字段
            "个人负担企额",      # OCR变体
        ])
    # 3) 兜底：个人现金支出（注意：医保结算单里"个人账户支出"也含"个人"和"支出"
    #    字样，因此必须用完整字段名，不能模糊匹配）
    if self_pay == 0:
        self_pay = _match_amount_after_label(text, ["个人现金支出"])
    if self_pay > 0:
        result["self_pay_amount"] = self_pay
    # 如果有报销金额但没有自费，自费=金额-报销
    if result["amount"] > 0 and result["insurance_amount"] > 0 and result["self_pay_amount"] == 0:
        result["self_pay_amount"] = result["amount"] - result["insurance_amount"]

    # ── 类型 ──
    if "住院" in text and ("一日清单" in text or "结算单" in text):
        result["receipt_type"] = "住院费"
    elif "门诊" in text:
        result["receipt_type"] = "门诊费"
    elif "CT" in text or "检查" in text:
        result["receipt_type"] = "检查费"
    elif "药" in text and "费" in text:
        result["receipt_type"] = "药费"
    else:
        # 从收据费目提取
        m = re.search(r'收据费目\s*\n?\s*(\S+?)\s*\n?\s*[\d.]', text)
        if m:
            result["receipt_type"] = m.group(1)

    # ── 住院天数 ──
    m = re.search(r'住院天数\s*\n?\s*(\d+)', text)
    if m:
        result["stay_days"] = int(m.group(1))
    elif result["receipt_type"] == "住院费":
        # 从入院-出院日期计算
        m_in = re.search(r'入院日期[：:]?\s*(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})', text)
        m_out = re.search(r'出院日期[：:]?\s*(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})', text)
        if m_in and m_out:
            from datetime import date
            try:
                d1 = date(int(m_in.group(1)), int(m_in.group(2)), int(m_in.group(3)))
                d2 = date(int(m_out.group(1)), int(m_out.group(2)), int(m_out.group(3)))
                result["stay_days"] = (d2 - d1).days
            except (ValueError, TypeError):
                pass

    # ── 票据尾号 ──
    m = re.search(r'票据号码[：:]\s*(\d+)', text)
    if m:
        result["receipt_tail"] = m.group(1)[-4:]  # 后4位
    else:
        m = re.search(r'单据号\s*\n?\s*(\d+)', text)
        if m:
            result["receipt_tail"] = m.group(1)[-4:]

    return result


def _is_medical_fee_item(item: dict) -> bool:
    """判断一个 fee_receipt 素材是否属于医疗费用（排除护理用品/住宿/交通等）。"""
    ocr_text = item.get("ocr_text", "") or ""
    title = item.get("title", "") or ""
    evidence_name = item.get("evidence_name", {})
    if isinstance(evidence_name, dict):
        original_filename = evidence_name.get("original_filename", "") or ""
        doc_type = evidence_name.get("doc_type", "") or ""
        doc_type_name = evidence_name.get("doc_type_name", "") or ""
    else:
        original_filename = ""
        doc_type = ""
        doc_type_name = ""

    # 排除关键词（标题/文件名层面）
    exclude_keywords = ["护理垫", "护理用品", "湿纸巾", "爽身粉", "护理床", "住宿", "酒店", "打车", "交通",
                        "高铁", "飞机", "地铁", "救护车", "房租", "日常用品", "案件交办", "补充材料清单",
                        "电子发票", "医疗器械发票", "收款收据"]
    for kw in exclude_keywords:
        if kw in title or kw in original_filename:
            return False

    # 如果原始文件名是"门诊费用*.jpg"→ 医疗
    if "门诊" in original_filename and ("费用" in original_filename or "收费" in original_filename):
        return True

    # 医院收费告知单 → 医疗
    if "收费告知" in title or "费用凭证" in ocr_text[:200]:
        return True

    # 医保结算单 → 医疗
    if "医保结算单" in title or "医保结算单" in ocr_text[:100]:
        return True

    # 住院一日清单 → 医疗
    if "住院" in title and "清单" in title:
        return True

    # 电子发票中含"医药"或"医疗器械" → 不是医疗费（是护理用品购买）
    if "电子发票" in title:
        return False

    # 收款收据（非医院收费） → 不是医疗费
    if title == "收款收据" and "收费告知" not in ocr_text[:300]:
        return False

    # 案件交办/材料清单等文档 → 不是医疗费
    if "案件交办" in title or "材料清单" in title or "登记表" in title:
        return False

    # 医院名称在OCR前500字出现 → 医疗
    if re.search(r'[\u4e00-\u9fa5]{2,8}(?:人民医院|中心医院|附属医院)', ocr_text[:500]):
        return True

    return False


def _gen_medical_fee_excel(fee_items: list[dict], comp_item: dict | None, case_name: str = "") -> bytes:
    """医疗费用汇总表 — 参考医疗费用汇总表（李明凤）.xls 格式。

    核心设计（与原版模板一致）：
    - 按费用类型汇总（不是逐条票据），通常只有 2-3 行数据：
      ① 门诊/检查费汇总行（所有门诊票据金额合计）
      ② 住院费汇总行（从医保结算单取总额、报销、自费、住院天数）
    - 同一医院名称合并单元格（名称列 B）
    - 住院费行：日期=入院-出院范围，住院天数从医保结算单提取
    - 门诊费行：日期取第一张门诊票据日期，票据尾号=费用凭证
    - 总计行：合并 B:C 写"总计"，汇总金额/报销/自费
    - 填充空行至 97 行
    - 仿宋字体，标题16号，内容12号，居中
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    COL_COUNT = 10
    headers = ["序号", "名称", "日期", "金额", "报销金额", "个人自费部分", "类型", "住院天数", "票据尾号", "证据页码"]

    # ── 标题行（行1，合并 A1:J1）──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=COL_COUNT)
    title_cell = ws.cell(row=1, column=1, value="医疗费用汇总表")
    title_cell.font = _FANGSONG_TITLE
    title_cell.alignment = _CENTER_ALIGN

    # ── 表头行（行2）──
    HEADER_ROW = 2
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=header)
        cell.font = _FANGSONG_HEADER
        cell.alignment = _CENTER_ALIGN
        cell.border = _THIN_BORDER

    # ════════════════════════════════════════════════════════════════════════
    # 第一步：从所有医疗票据中解析数据，按"类型"汇总
    # 原版模板逻辑：
    # - 住院费行：从医保结算单取总额/报销/自费/住院天数
    # - 门诊/检查费行：只有有统筹支付（报销）的门诊票据才算（如CT检查费140元）
    #   那些没有统筹支付的小额床位费（10/20元）已包含在住院费总额中，不单独列
    # ════════════════════════════════════════════════════════════════════════
    hospital_name = ""
    outpatient_total = 0.0        # 门诊费用合计（仅有统筹支付的）
    outpatient_insurance = 0.0    # 门诊报销合计
    outpatient_self_pay = 0.0     # 门诊自费合计
    outpatient_first_date = ""    # 最早门诊日期

    inpatient_amount = 0.0        # 住院费总额
    inpatient_insurance = 0.0     # 住院报销
    inpatient_self_pay = 0.0      # 住院自费
    inpatient_date_range = ""     # 入院-出院日期
    inpatient_stay_days = ""      # 住院天数

    for item in fee_items:
        if not _is_medical_fee_item(item):
            continue

        ocr_text = item.get("ocr_text", "") or ""
        parsed = _parse_medical_fee_from_ocr(ocr_text)

        # 提取医院名称（取第一个匹配的）
        if not hospital_name and parsed["hospital_name"]:
            hospital_name = parsed["hospital_name"]

        receipt_type = parsed["receipt_type"]
        amount = parsed["amount"] or _to_float(item.get("amount", 0))
        # Fallback: 如果OCR没解析出金额，从结构化 fees.total_amount 取
        if amount == 0:
            fees_data = item.get("fees", {}) or item.get("fee_detail", {})
            if isinstance(fees_data, dict):
                amt = fees_data.get("total_amount", 0)
                if amt and amt > 0:
                    amount = _to_float(amt)
        insurance = parsed["insurance_amount"]
        # Fallback: 从结构化数据取报销金额
        if insurance == 0:
            fees_data = item.get("fees", {}) or item.get("fee_detail", {})
            if isinstance(fees_data, dict):
                ins = fees_data.get("insurance_amount", 0)
                if ins:
                    insurance = _to_float(ins)
        self_pay = parsed["self_pay_amount"]
        # Fallback: 从结构化数据取自费金额
        if self_pay == 0:
            fees_data = item.get("fees", {}) or item.get("fee_detail", {})
            if isinstance(fees_data, dict):
                oop = fees_data.get("out_of_pocket", 0)
                if oop:
                    self_pay = _to_float(oop)
        if amount > 0 and insurance > 0 and self_pay == 0:
            self_pay = amount - insurance

        if receipt_type == "住院费":
            # 住院费：从医保结算单取数据
            # 选择策略：优先取"有报销金额"的那张（字段完整、含结算信息）
            # 其次取"金额最大"的那张
            should_pick = False
            if insurance > 0 and inpatient_insurance == 0:
                # 当前这张有报销，之前那张没有 → 取当前这张
                should_pick = True
            elif insurance > 0 and amount > inpatient_amount:
                # 当前这张有报销且金额更大 → 取当前这张
                should_pick = True
            elif inpatient_amount == 0:
                # 之前一张金额为0（未取到） → 取当前这张
                should_pick = True
            if should_pick:
                inpatient_amount = amount
                inpatient_insurance = insurance
                inpatient_self_pay = self_pay
            # 住院天数和日期范围
            if parsed["stay_days"]:
                inpatient_stay_days = parsed["stay_days"]
            if parsed["date"]:
                inpatient_date_range = parsed["date"]
            # 如果OCR没取到日期，从结构化数据取
            if not inpatient_date_range:
                evidence_name = item.get("evidence_name", {})
                if isinstance(evidence_name, dict):
                    d = evidence_name.get("date", "")
                    if d:
                        inpatient_date_range = d
        elif insurance > 0 or receipt_type != "住院费":
            # 门诊/检查费：只保留金额最大的那张（通常是CT/检查费，其它小票已含在住院费总额中）
            # 原版模板：检查费140元 + 报销31.5 + 自费108.5 (仅此一张，不含床位费小票)
            if amount > outpatient_total:
                outpatient_total = amount
                outpatient_insurance = insurance
                outpatient_self_pay = self_pay
                d = parsed["date"]
                if d:
                    outpatient_first_date = d

    # 门诊票据尾号统一写"费用凭证"（原版模板如此）
    outpatient_receipt_tail = "费用凭证" if outpatient_total > 0 else ""

    # 如果没有医院名，尝试从任意OCR文本提取
    if not hospital_name:
        for item in fee_items:
            if not _is_medical_fee_item(item):
                continue
            ocr_text = item.get("ocr_text", "") or ""
            parsed = _parse_medical_fee_from_ocr(ocr_text)
            if parsed["hospital_name"]:
                hospital_name = parsed["hospital_name"]
                break

    # ════════════════════════════════════════════════════════════════════════
    # 第二步：按类型写入汇总行（门诊费 + 住院费）
    # ════════════════════════════════════════════════════════════════════════
    row_idx = HEADER_ROW + 1
    seq = 1
    total_amount = 0.0
    total_insurance = 0.0
    total_self_pay = 0.0
    total_stay_days = ""

    summary_rows: list[dict] = []

    # 门诊/检查费汇总行
    if outpatient_total > 0:
        # 判断类型名称
        outpatient_type = "检查费" if outpatient_insurance > 0 else "门诊费"
        summary_rows.append({
            "seq": seq,
            "name": hospital_name,
            "date": outpatient_first_date if outpatient_first_date else "",
            "amount": outpatient_total,
            "insurance": outpatient_insurance,
            "self_pay": outpatient_self_pay,
            "type": outpatient_type,
            "stay_days": "",
            "receipt_tail": outpatient_receipt_tail,
            "evidence_page": "",
        })
        total_amount += outpatient_total
        total_insurance += outpatient_insurance
        total_self_pay += outpatient_self_pay
        seq += 1

    # 住院费汇总行
    if inpatient_amount > 0:
        summary_rows.append({
            "seq": seq,
            "name": hospital_name,
            "date": inpatient_date_range,
            "amount": inpatient_amount,
            "insurance": inpatient_insurance,
            "self_pay": inpatient_self_pay,
            "type": "住院费",
            "stay_days": inpatient_stay_days,
            "receipt_tail": "医保结算单",
            "evidence_page": "",
        })
        total_amount += inpatient_amount
        total_insurance += inpatient_insurance
        total_self_pay += inpatient_self_pay
        total_stay_days = inpatient_stay_days
        seq += 1

    # 写入汇总行
    first_data_row = row_idx
    for row_data in summary_rows:
        ws.cell(row=row_idx, column=1, value=row_data["seq"])
        ws.cell(row=row_idx, column=2, value=row_data["name"])
        ws.cell(row=row_idx, column=3, value=row_data["date"])
        ws.cell(row=row_idx, column=4, value=row_data["amount"] if row_data["amount"] else "")
        ws.cell(row=row_idx, column=5, value=row_data["insurance"] if row_data["insurance"] else "")
        ws.cell(row=row_idx, column=6, value=row_data["self_pay"] if row_data["self_pay"] else "")
        ws.cell(row=row_idx, column=7, value=row_data["type"])
        ws.cell(row=row_idx, column=8, value=row_data["stay_days"] if row_data["stay_days"] else "")
        ws.cell(row=row_idx, column=9, value=row_data["receipt_tail"])
        ws.cell(row=row_idx, column=10, value=row_data["evidence_page"])

        for col in range(1, COL_COUNT + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = _FANGSONG_CELL
            cell.border = _THIN_BORDER
            cell.alignment = _CENTER_ALIGN
            if col in (4, 5, 6) and cell.value and isinstance(cell.value, (int, float)):
                cell.number_format = _MONEY_FORMAT

        row_idx += 1

    # 合并名称列（同一医院）
    if row_idx > first_data_row + 1:
        ws.merge_cells(start_row=first_data_row, start_column=2, end_row=row_idx - 1, end_column=2)

    # ── 总计行 ──
    total_row = row_idx
    ws.merge_cells(start_row=total_row, start_column=2, end_row=total_row, end_column=3)
    ws.cell(row=total_row, column=1, value="")
    ws.cell(row=total_row, column=2, value="总计")
    ws.cell(row=total_row, column=4, value=total_amount)
    ws.cell(row=total_row, column=5, value=total_insurance if total_insurance else "")
    ws.cell(row=total_row, column=6, value=total_self_pay if total_self_pay else "")
    ws.cell(row=total_row, column=8, value=total_stay_days if total_stay_days else "")
    for col in range(1, COL_COUNT + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.font = _FANGSONG_BOLD
        cell.border = _THIN_BORDER
        cell.fill = _TOTAL_FILL
        cell.alignment = _CENTER_ALIGN
        if col in (4, 5, 6) and cell.value and isinstance(cell.value, (int, float)):
            cell.number_format = _MONEY_FORMAT
    row_idx = total_row + 1

    # ── 填充空行到约 97 行 ──
    TARGET_ROWS = 97
    while row_idx <= TARGET_ROWS:
        for col in range(1, COL_COUNT + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = _THIN_BORDER
            cell.font = _FANGSONG_CELL
        row_idx += 1

    # ── 列宽 ──
    col_widths = [8, 28, 22, 14, 14, 16, 10, 10, 12, 10]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _gen_nursing_supplies_excel(fee_items: list[dict], comp_item: dict | None, case_name: str = "") -> bytes:
    """医护用品费统计表 — 参考护理用品费统计表（李明凤）.xls 格式。

    表头：序号 | 名称 | 日期 | 金额 | 类型 | 票据尾号
    合计行："合计" | | | 总金额 | |
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    COL_COUNT = 6
    headers = ["序号", "名称", "日期", "金额", "类型", "票据尾号"]

    _write_title_row(ws, "医护用品费统计表", COL_COUNT)
    _write_header_row(ws, headers, 2)

    row_idx = 3
    seq = 1
    total = 0.0

    # 筛选护理用品类
    for item in fee_items:
        ft = item.get("fee_type", "")
        if ft != "nursing_supplies" and "护理用品" not in ft and "医护用品" not in ft:
            continue
        amount = item.get("amount", 0) or 0
        total += amount
        name = item.get("title", "") or item.get("hospital_name", "")
        ws.cell(row=row_idx, column=1, value=seq)
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value=item.get("date", ""))
        ws.cell(row=row_idx, column=4, value=amount if amount else "")
        ws.cell(row=row_idx, column=5, value=item.get("receipt_type", ""))
        ws.cell(row=row_idx, column=6, value=item.get("receipt_tail", ""))
        _apply_cell_style(ws, row_idx, COL_COUNT, money_cols={4})
        row_idx += 1
        seq += 1

    # 合计行
    ws.cell(row=row_idx, column=1, value="合计")
    ws.cell(row=row_idx, column=1).font = _BOLD_FONT
    ws.cell(row=row_idx, column=1).alignment = _CENTER_ALIGN
    ws.cell(row=row_idx, column=4, value=total)
    ws.cell(row=row_idx, column=4).font = _BOLD_FONT
    ws.cell(row=row_idx, column=4).number_format = _MONEY_FORMAT
    for col in range(1, COL_COUNT + 1):
        ws.cell(row=row_idx, column=col).border = _THIN_BORDER
        ws.cell(row=row_idx, column=col).fill = _TOTAL_FILL

    col_widths = [8, 30, 18, 14, 10, 14]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _gen_accommodation_excel(fee_items: list[dict], comp_item: dict | None, case_name: str = "") -> bytes:
    """住宿费统计表 — 参考住宿费统计表（李明凤）.xls 格式。

    表头：序号 | 类型 | 住宿人员 | 日期 | 金额 | 备注 | 票据尾号
    总计行
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    COL_COUNT = 7
    headers = ["序号", "类型", "住宿人员", "日期", "金额", "备注", "票据尾号"]

    _write_title_row(ws, "住宿费统计表", COL_COUNT)
    _write_header_row(ws, headers, 2)

    row_idx = 3
    seq = 1
    total = 0.0

    for item in fee_items:
        ft = item.get("fee_type", "")
        if ft and "住宿" not in ft and ft != "accommodation_fee":
            continue
        amount = item.get("amount", 0) or 0
        total += amount
        ws.cell(row=row_idx, column=1, value=seq)
        ws.cell(row=row_idx, column=2, value=item.get("receipt_type", "") or "住宿费")
        ws.cell(row=row_idx, column=3, value="")
        ws.cell(row=row_idx, column=4, value=item.get("date", ""))
        ws.cell(row=row_idx, column=5, value=amount if amount else "")
        ws.cell(row=row_idx, column=6, value=item.get("title", ""))
        ws.cell(row=row_idx, column=7, value=item.get("receipt_tail", ""))
        _apply_cell_style(ws, row_idx, COL_COUNT, money_cols={5})
        row_idx += 1
        seq += 1

    # 总计行
    ws.cell(row=row_idx, column=2, value="总计")
    ws.cell(row=row_idx, column=2).font = _BOLD_FONT
    ws.cell(row=row_idx, column=2).alignment = _CENTER_ALIGN
    ws.cell(row=row_idx, column=5, value=total)
    ws.cell(row=row_idx, column=5).font = _BOLD_FONT
    ws.cell(row=row_idx, column=5).number_format = _MONEY_FORMAT
    for col in range(1, COL_COUNT + 1):
        ws.cell(row=row_idx, column=col).border = _THIN_BORDER
        ws.cell(row=row_idx, column=col).fill = _TOTAL_FILL

    col_widths = [8, 12, 14, 18, 14, 20, 14]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _gen_transport_excel(fee_items: list[dict], comp_item: dict | None, case_name: str = "") -> bytes:
    """交通费统计表。

    表头：序号 | 名称 | 日期 | 金额 | 备注 | 票据尾号
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    COL_COUNT = 6
    headers = ["序号", "名称", "日期", "金额", "备注", "票据尾号"]

    _write_title_row(ws, "交通费统计表", COL_COUNT)
    _write_header_row(ws, headers, 2)

    row_idx = 3
    seq = 1
    total = 0.0

    for item in fee_items:
        ft = item.get("fee_type", "")
        if ft and "交通" not in ft and ft != "transport_fee" and ft != "transportation":
            continue
        amount = item.get("amount", 0) or 0
        total += amount
        name = item.get("title", "") or item.get("hospital_name", "")
        ws.cell(row=row_idx, column=1, value=seq)
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value=item.get("date", ""))
        ws.cell(row=row_idx, column=4, value=amount if amount else "")
        ws.cell(row=row_idx, column=5, value="")
        ws.cell(row=row_idx, column=6, value=item.get("receipt_tail", ""))
        _apply_cell_style(ws, row_idx, COL_COUNT, money_cols={4})
        row_idx += 1
        seq += 1

    # 总计行
    ws.cell(row=row_idx, column=2, value="总计")
    ws.cell(row=row_idx, column=2).font = _BOLD_FONT
    ws.cell(row=row_idx, column=2).alignment = _CENTER_ALIGN
    ws.cell(row=row_idx, column=4, value=total)
    ws.cell(row=row_idx, column=4).font = _BOLD_FONT
    ws.cell(row=row_idx, column=4).number_format = _MONEY_FORMAT
    for col in range(1, COL_COUNT + 1):
        ws.cell(row=row_idx, column=col).border = _THIN_BORDER
        ws.cell(row=row_idx, column=col).fill = _TOTAL_FILL

    col_widths = [8, 28, 18, 14, 20, 14]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _gen_simple_fee_excel(title: str, comp_item: dict | None, calc_basis: str = "") -> bytes:
    """简单格式 Excel（误工费/护理费/住院伙食补助费/营养费/残疾赔偿金/死亡赔偿金/鉴定费/精神损害抚慰金/被扶养人生活费）。

    表头：序号 | 项目 | 金额 | 计算依据
    数据行 + 合计行
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    COL_COUNT = 4
    headers = ["序号", "项目", "金额", "计算依据"]

    _write_title_row(ws, title, COL_COUNT)
    _write_header_row(ws, headers, 2)

    row_idx = 3

    # 从 comp_item 取金额和计算依据
    amount = 0.0
    basis = calc_basis
    if comp_item:
        amount = _to_float(comp_item.get("manual_amount") or comp_item.get("amount", 0))
        basis = comp_item.get("calculation_basis", "") or calc_basis

    ws.cell(row=row_idx, column=1, value=1)
    ws.cell(row=row_idx, column=2, value=title)
    ws.cell(row=row_idx, column=3, value=amount if amount else "")
    ws.cell(row=row_idx, column=4, value=basis)
    _apply_cell_style(ws, row_idx, COL_COUNT, money_cols={3})
    row_idx += 1

    # 合计行
    ws.cell(row=row_idx, column=2, value="合计")
    ws.cell(row=row_idx, column=2).font = _BOLD_FONT
    ws.cell(row=row_idx, column=2).alignment = _CENTER_ALIGN
    ws.cell(row=row_idx, column=3, value=amount if amount else "")
    ws.cell(row=row_idx, column=3).font = _BOLD_FONT
    ws.cell(row=row_idx, column=3).number_format = _MONEY_FORMAT
    for col in range(1, COL_COUNT + 1):
        ws.cell(row=row_idx, column=col).border = _THIN_BORDER
        ws.cell(row=row_idx, column=col).fill = _TOTAL_FILL

    col_widths = [8, 22, 16, 55]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_fee_excel_zip(
    catalog_data: dict,
    analysis_result: dict,
    compensation_items: list[dict] | None,
) -> bytes:
    """生成 10 项赔偿明细 Excel ZIP 包。

    Args:
        catalog_data: 案件清单数据（含 groups, fee_summary）
        analysis_result: 分析结果
        compensation_items: 前端传来的当前编辑金额列表 [{fee_type, amount/manual_amount, ...}]

    Returns:
        ZIP 字节流，包含 10 个 xlsx 文件
    """
    case_name = analysis_result.get("case_name", "") or ""
    case_type = analysis_result.get("case_type", "injury")

    # 构建 fee_type → comp_item 映射
    comp_map: dict[str, dict] = {}
    if compensation_items:
        for item in compensation_items:
            ft = item.get("fee_type", "")
            if ft:
                comp_map[ft] = item

    # 从 catalog_data 提取费用票据
    fee_items = _collect_fee_receipt_items_v2(catalog_data)

    # 构建 ZIP
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        # 1. 医疗费
        medical_bytes = _gen_medical_fee_excel(
            fee_items, comp_map.get("medical_fee"), case_name
        )
        zf.writestr("01_医疗费.xlsx", medical_bytes)

        # 2. 误工费
        lost_wages_bytes = _gen_simple_fee_excel(
            "误工费", comp_map.get("lost_wages"), "误工收入 × 误工天数（参照鉴定意见）"
        )
        zf.writestr("02_误工费.xlsx", lost_wages_bytes)

        # 3. 护理费 — 包含医护用品统计表
        nursing_main = _gen_simple_fee_excel(
            "护理费", comp_map.get("nursing_fee"), "住院期间护理费 = 护理标准 × 住院天数"
        )
        zf.writestr("03_护理费.xlsx", nursing_main)

        # 护理用品费统计表（作为护理费的补充明细）
        nursing_supplies_bytes = _gen_nursing_supplies_excel(
            fee_items, comp_map.get("nursing_fee"), case_name
        )
        zf.writestr("03_护理费_医护用品费统计表.xlsx", nursing_supplies_bytes)

        # 4. 住院伙食补助费
        food_bytes = _gen_simple_fee_excel(
            "住院伙食补助费", comp_map.get("food_subsidy"), "住院伙食补助标准 × 住院天数"
        )
        zf.writestr("04_住院伙食补助费.xlsx", food_bytes)

        # 5. 营养费
        nutrition_bytes = _gen_simple_fee_excel(
            "营养费", comp_map.get("nutrition_fee"), "营养费标准 × 营养期天数（参照鉴定意见）"
        )
        zf.writestr("05_营养费.xlsx", nutrition_bytes)

        # 6. 残疾赔偿金 / 死亡赔偿金（根据案件类型只生成一个）
        if case_type == "death":
            death_bytes = _gen_simple_fee_excel(
                "死亡赔偿金", comp_map.get("death_compensation"),
                "受诉法院所在地上年度城镇居民人均可支配收入 × 20年",
            )
            zf.writestr("06_死亡赔偿金.xlsx", death_bytes)
        else:
            disability_bytes = _gen_simple_fee_excel(
                "残疾赔偿金", comp_map.get("disability_compensation"),
                "受诉法院所在地上年度城镇居民人均可支配收入 × 20年 × 伤残系数",
            )
            zf.writestr("06_残疾赔偿金.xlsx", disability_bytes)

        # 7. 交通住宿费 — 包含交通费和住宿费两个统计表
        transport_bytes = _gen_transport_excel(
            fee_items, comp_map.get("transport_fee"), case_name
        )
        zf.writestr("07_交通费统计表.xlsx", transport_bytes)

        accommodation_bytes = _gen_accommodation_excel(
            fee_items, comp_map.get("transport_fee"), case_name
        )
        zf.writestr("07_住宿费统计表.xlsx", accommodation_bytes)

        # 8. 鉴定费
        appraisal_bytes = _gen_simple_fee_excel(
            "鉴定费", comp_map.get("appraisal_fee"), "凭票据实报实销"
        )
        zf.writestr("08_鉴定费.xlsx", appraisal_bytes)

        # 9. 精神损害抚慰金
        spiritual_bytes = _gen_simple_fee_excel(
            "精神损害抚慰金", comp_map.get("spiritual_damage"),
            "根据伤残等级/死亡后果酌定",
        )
        zf.writestr("09_精神损害抚慰金.xlsx", spiritual_bytes)

        # 10. 被扶养人生活费（如有）
        dependent_item = comp_map.get("dependent_living")
        dependent_bytes = _gen_simple_fee_excel(
            "被扶养人生活费", dependent_item,
            "受诉法院所在地上年度城镇居民人均消费支出 × 扶养年限",
        )
        zf.writestr("10_被扶养人生活费.xlsx", dependent_bytes)

    return zip_buffer.getvalue()
