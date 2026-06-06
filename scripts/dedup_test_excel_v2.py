"""Regenerate test report Excel: deduplicate with docstring-enriched content."""
import ast, os, json, pandas as pd
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load test dataset
with open('docs/test_dataset_2026-06-06.json', 'r', encoding='utf-8') as f:
    tests = json.load(f)

# Load docstrings
with open('docs/test_docstrings.json', 'r', encoding='utf-8') as f:
    docstrings = json.load(f)

module_map = {
    'test_api.py': 'API 健康检查与管理',
    'test_bailian_ocr.py': 'OCR 引擎(百炼)',
    'test_callback.py': '业务回调',
    'test_cross_page_merger.py': '跨页段落合并',
    'test_e2e_workflows.py': '端到端工作流',
    'test_evidence.py': '证据分类/目录/分析/导出',
    'test_header_footer_cleaner.py': '页眉页脚清除',
    'test_heading_parser.py': '标题层级解析',
    'test_image_enhancer.py': '图像增强',
    'test_json_exporter.py': 'JSON 导出',
    'test_layout_detector.py': '版面检测',
    'test_list_detector.py': '列表检测',
    'test_minio_client.py': 'MinIO 存储客户端',
    'test_ocr_batch_processor.py': 'OCR 批处理',
    'test_ocr_engine.py': 'OCR 引擎(PaddleOCR)',
    'test_orientation_standalone.py': '页面方向检测',
    'test_paragraph_grouper.py': '段落分组',
    'test_pdf_classifier.py': 'PDF 分类器',
    'test_pdf_splitter.py': 'PDF 页面拆分',
    'test_quality_scorer.py': '质量评分',
    'test_scan_api.py': '扫描 API',
    'test_stream_publisher.py': 'SSE 流式推送',
    'test_table_recognizer.py': '表格识别',
    'test_text_pdf_extractor.py': '文本 PDF 提取',
    'test_validator.py': 'PDF 校验器',
    'test_watcher.py': '文件监控',
}


def humanize_test_name(name):
    """Convert test_name to human-readable Chinese description."""
    n = name.replace('test_', '')
    # Comprehensive mapping of test name patterns to Chinese descriptions
    # Format: (substring_match, description)
    mapping = [
        # API
        ('ping_returns_pong', 'Ping返回pong'),
        ('health_when_all_services_up', '所有服务正常时health返回ok'),
        ('health_when_db_fails', '数据库故障时health返回db=error'),
        ('health_when_redis_fails', 'Redis故障时health返回redis=error'),
        ('health_when_minio_fails', 'MinIO故障时health返回minio=error'),
        ('health_ping_always_public', 'health/ping端点无需API Key即可访问'),
        ('admin_stats_returns_expected_schema', 'admin/stats返回正确schema'),
        ('admin_stats_no_tasks', '无任务时admin/stats返回零值'),
        ('admin_queue_returns_tasks', 'admin/queue返回排队任务列表'),
        ('admin_queue_empty', '队列为空时返回空列表'),
        ('admin_without_key_rejected', '无API Key访问/admin被拒绝'),
        ('admin_with_wrong_key_rejected', '错误API Key返回403'),
        ('admin_with_valid_key_accepted', '正确API Key通过验证'),
        ('admin_with_lowercase_header', '小写x-api-key头也被接受'),
        # Bailian OCR
        ('factory_returns_bailian_when_configured', '配置bailian时返回BailianOCREngine'),
        ('bailian_engine_has_required_interface', 'BailianOCREngine实现OCREngine接口'),
        ('factory_get_ocr_engine', 'get_ocr_engine()按配置返回引擎'),
        ('factory_default_is_paddle', '默认ocr_engine_type=paddle'),
        ('real_api_recognize_text', '真实百炼API识别文字'),
        ('real_api_structured_json', '百炼API结构化JSON输出'),
        ('real_api_save_result', '百炼API+save_result写入文件'),
        ('parse_valid_json_array', '解析合法JSON数组'),
        ('parse_markdown_code_block', '解析Markdown代码块包裹的JSON'),
        ('parse_fallback_plain_text', '回退解析纯文本'),
        ('parse_empty_response', '处理空响应'),
        ('parse_partial_json_in_text', '从混杂文本提取JSON数组'),
        ('output_format_matches_ocr_engine', '百炼输出格式与OCREngine一致'),
        ('batch_processor_accepts_bailian_engine', 'batch_processor接受百炼引擎'),
        ('engine_not_ready_without_api_key', '无API Key时引擎不可用但不崩溃'),
        ('base64_encoding_png', 'PNG图片编码为JPEG Base64'),
        ('base64_encoding_jpeg', 'JPEG图片正确Base64编码'),
        ('base64_invalid_path', '无效路径抛出FileNotFoundError'),
        ('real_api_ocr', '真实百炼API OCR调用'),
        # Callback
        ('callback_success', '有效回调处理成功'),
        ('callback_invalid_data', '无效回调数据返回422'),
        ('callback_timeout_retry', '回调超时后重试'),
        ('callback_not_found', 'task不存在回调返回404'),
        # Cross page merger
        ('continuation_dash', '破折号续接跨页段落合并'),
        ('continuation_ellipsis', '省略号续接跨页段落合并'),
        ('no_continuation', '不同段落跨页不合并'),
        ('simple_merge', '简单跨页续接合并'),
        ('empty_text', '空文本不触发合并'),
        ('empty_pages', '空页面正常处理'),
        ('single_page_no_merge', '单页文档无合并操作'),
        ('no_merge_needed', '无需合并时不产生修改'),
        ('flat_merge', '无缩进层级的多页合并'),
        ('no_boundary_no_change', '无跨页边界时不修改内容'),
        # Evidence
        ('identity_keyword_match', '关键词"身份证"匹配identity_id_card'),
        ('identity_business_license', '"营业执照"匹配identity_defendant'),
        ('identity_credit_code', '"统一社会信用代码"匹配identity_defendant'),
        ('medical_keyword_match', '关键词"病历"匹配medical_record'),
        ('fee_keyword_match', '关键词"医疗费发票"匹配fee_receipt'),
        ('death_cert_keyword_match', '关键词"死亡证明"匹配death_certificate'),
        ('appraisal_keyword_match', '关键词"鉴定意见"匹配appraisal'),
        ('multiple_keywords_higher_confidence', '多关键词匹配置信度≥0.6'),
        ('all_categories_in_order', 'CATEGORY_ORDER包含全部9个分类'),
        ('empty_text_returns_other', '空文本分类为other_evidence'),
        ('default_case_type_is_injury', '默认案件类型=injury'),
        ('llm_fallback_called_when_no_keyword_match', '关键词不匹配时调用LLM分类'),
        ('llm_returns_valid_json', 'LLM返回有效JSON正确解析'),
        ('llm_returns_no_json', 'LLM返回非JSON时回退处理'),
        ('llm_death_case_excludes_death_cert_for_injury', 'injury案件LLM prompt标注death_certificate限制'),
        # Fee
        ('single_fee_item', '单项费用(医疗费12345.67)正确汇总'),
        ('multiple_same_fee_type', '同类费用累加(医疗费1000+2000=3000)'),
        ('different_fee_types', '不同费用类型分别统计'),
        ('zero_amount_excluded', '金额=0的费用项排除'),
        ('negative_amount_excluded', '金额<0的费用项排除'),
        ('non_numeric_amount_ignored', '非数字金额忽略'),
        ('no_fee_type_ignored', '缺fee_type的项忽略'),
        ('empty_fee_detail', '空fee_detail{}忽略'),
        ('none_fee_detail', 'fee_detail=None忽略'),
        ('empty_materials', '材料列表为空返回{}'),
        ('rounding_to_2_decimals', '金额四舍五入到2位小数'),
        ('integer_amount', '整数金额正确处理'),
        # Title/Purpose
        ('identity_title', 'identity_id_card标题含"身份证"'),
        ('other_evidence_just_filename', 'other_evidence标题含"其他证据"'),
        ('no_filename', '无文件名时标题使用默认格式'),
        ('identity_purpose', 'identity_id_card用途含"主体身份"'),
        # Schema
        ('case_response_matches_model', 'Response Schema覆盖Model关键字段'),
        ('create_case_request_matches', 'Request Schema与创建参数一致'),
        ('total_endpoint_count', 'API端点数≥20'),
        ('frontend_types_match_backend_schemas', '前端类型与后端Schema核心字段一致'),
        ('router_registered', 'evidence router已注册到app'),
        ('category_names', 'CATEGORY_NAMES映射完整'),
        # Services
        ('services_init_importable', 'services.evidence包可正常导入'),
        ('export_bundle_task_calls_correct_service', 'export_evidence_bundle任务存在可调用'),
        ('ocr_pipeline_uses_correct_ocr_service', '_run_ocr_pipeline使用OCR服务'),
        # Static analysis
        ('match_route', '路由定义正则匹配'),
        ('check_params', '端点参数定义检查'),
        ('check_response', '端点返回类型检查'),
        ('check_imports', '模块导入链完整性检查'),
        # Bundle
        ('zip_creation_basic', '基础ZIP打包创建'),
        ('create_bundle_success', '5个文件打包成功'),
        ('bundle_missing_file', '打包时1个文件缺失处理'),
        ('empty_zip', '空列表打包返回空ZIP'),
        # Excel
        ('excel_headers', 'Excel表头字段正确'),
        ('excel_data_rows', '数据行内容正确'),
        # Header footer cleaner
        ('page_numbers_removed', '重复页脚文本移除'),
        ('headers_removed', '重复页眉文本移除'),
        ('mixed_content_preserved', '正文内容不被误删'),
        ('text_similarity_identical', '相同文本相似度=1.0'),
        ('text_similarity_different', '不同文本相似度<0.5'),
        ('empty', '空输入正常处理'),
        ('few_pages_no_cleanup', '少于min_pages的文档不清理'),
        ('position_is_header', '顶部区域文本识别为页眉'),
        ('position_is_footer', '底部区域文本识别为页脚'),
        # Heading parser
        ('level_1_heading', '"一、"识别为一级标题'),
        ('level_2_heading', '"（一）"识别为二级标题'),
        ('level_3_heading', '"1."识别为三级标题'),
        ('level_4_heading', '"(1)"识别为四级标题'),
        ('level_5_heading', '"①"识别为五级标题'),
        ('no_heading_text', '普通文本不识别为标题'),
        ('mixed_heading_and_text', '混合标题+段落正确分层'),
        ('register_custom_pattern', '自定义标题正则注册'),
        ('registry_contains_defaults', '注册表含默认匹配模式'),
        ('parse_headings_empty_blocks', '空block列表返回空结果'),
        ('parse_headings_no_headings', '无标题block返回平铺段落'),
        # Image enhancer
        ('auto_level_stretch', '直方图拉伸到[0,255]'),
        ('deskew_detect_angle', '倾斜角度检测'),
        ('crop_black_border', '黑边裁剪'),
        ('crop_no_border_preserves_size', '无黑边时尺寸不变'),
        ('default_settings', '默认增强参数'),
        ('enhance_basic', '基础增强流程'),
        ('init_params', 'ImageEnhancer初始化参数'),
        # JSON exporter
        ('export_basic_result', '基础JSON导出'),
        ('export_with_table', '含表格的JSON导出'),
        ('export_empty_result', '空结果导出'),
        # Layout detector
        ('detect_single_column', '单栏版面检测'),
        ('detect_multi_column', '多栏版面检测'),
        ('detect_mixed_layout', '混合版面检测'),
        ('detect_columns', '分栏检测'),
        ('is_table_row', '表格行判断'),
        ('empty_bbox', '空bbox处理'),
        ('no_overlap', '无重叠矩形'),
        ('bbox_to_rect', 'BBox→Rect坐标转换'),
        ('rect_overlap', '矩形重叠IoU计算'),
        # List detector
        ('detect_ordered_list', '有序列表检测'),
        ('detect_unordered_list', '无序列表检测'),
        ('match_list_item', '列表项匹配'),
        ('detect_none_on_ambiguous', '模糊文本返回None'),
        ('empty_input', '空输入返回空列表'),
        ('empty_text', '空文本返回空列表'),
        ('empty_blocks', '空block列表返回空'),
        # MinIO
        ('upload_small_file', '小文件上传'),
        ('upload_large_file', '大文件上传'),
        ('upload_success', '上传成功'),
        ('upload_empty_file', '空文件上传'),
        ('download_success', '下载成功'),
        ('download_not_found', '下载不存在对象静默/404'),
        ('delete_success', '删除已存在对象'),
        ('delete_not_found', '删除不存在对象静默成功'),
        ('presigned_get_url', 'GET预签名URL生成'),
        ('presigned_put_url', 'PUT预签名URL生成'),
        ('get_presigned_url_default_expiry', '预签名URL默认有效期'),
        ('bucket_exists', 'Bucket存在性检查'),
        ('create_bucket', '创建新Bucket'),
        ('ensure_buckets_creates_missing', '自动创建缺失的Bucket'),
        ('delete_task_objects', '批量删除任务对象'),
        ('object_exists', '对象存在性检查'),
        ('global_singleton', '全局单例模式验证'),
        # OCR batch
        ('batch_sequential', '顺序批处理'),
        ('batch_concurrent', '并发批处理'),
        ('batch_with_failure', '含失败项批处理'),
        ('batch_empty', '空列表批处理'),
        ('empty_image_list', '空图片列表批处理'),
        # OCR engine
        ('init_cpu_mode', 'CPU模式初始化'),
        ('load_model', '预加载OCR模型'),
        ('recognize_single', '单图OCR识别'),
        ('recognize_batch', '批量OCR识别'),
        ('recognize_not_loaded_returns_empty', '模型未加载返回空结果'),
        ('recognize_empty_page', '空页面返回空结果'),
        ('recognize_skips_empty_text', '空文本页面跳过'),
        ('save_result', 'OCR结果保存到文件'),
        ('default_is_paddle', '默认引擎类型=paddle'),
        ('default_not_ready', '默认状态未就绪'),
        ('load_easyocr_fallback_when_paddle_missing', 'PaddleOCR缺失时回退EasyOCR'),
        # Orientation
        ('detect_0_degrees', '0度(正常)方向检测'),
        ('detect_90_degrees', '90度旋转检测'),
        ('detect_180_degrees', '180度旋转检测'),
        ('detect_270_degrees', '270度旋转检测'),
        # Paragraph grouper
        ('split_paragraphs', '连续文本段落拆分'),
        ('group_paragraphs', '含缩进段落分组'),
        ('build_hierarchy', '3级缩进层级构建'),
        ('page_number_detection', '页码文本检测排除'),
        ('empty_text', '空文本处理'),
        ('empty_blocks', '空block处理'),
        # PDF classifier
        ('classify_text_pdf', '文本PDF分类'),
        ('classify_scanned_pdf', '扫描PDF分类'),
        ('classify_empty_pdf', '空PDF分类'),
        ('classify_no_fitz_fallback', 'fitz不可用时回退'),
        ('pdf_info', 'PDF元信息提取'),
        # PDF splitter
        ('default_dpi', '默认DPI=200'),
        ('default_dpi_from_settings_when_none_passed', '未传DPI时从settings取默认值'),
        ('split_all_pages', '全部页面拆分'),
        ('split_page_range', '指定页码范围拆分'),
        ('split_end_page_beyond_total', '结束页超出总页数处理'),
        ('split_output_filenames', '输出文件名格式验证'),
        ('dpi_zoom_calculation', 'DPI缩放矩阵计算'),
        ('split_to_bytes', '拆分为字节流'),
        ('init_params', 'PDFSplitter初始化参数'),
        # Quality scorer
        ('ocr_confidence_score', 'OCR置信度评分'),
        ('heading_quality_score', '标题质量评分'),
        ('max_heading_depth', '最大标题深度'),
        ('structure_completeness', '结构完整性评分'),
        ('score_structure', '基础质量评分'),
        ('empty_ocr', '空OCR结果评分'),
        ('no_structure', '无结构文档评分'),
        ('empty_sections', '空章节评分'),
        ('default_params', '默认评分参数'),
        # Scan API
        ('upload_success', 'PDF上传成功'),
        ('upload_empty_file', '空文件上传'),
        ('upload_exceed_max_pages', '超过最大页数上传'),
        ('upload_invalid_type', '非PDF文件上传'),
        ('upload_audio_file', '音频文件上传'),
        ('list_tasks', '任务列表查询'),
        ('list_empty', '空任务列表'),
        ('list_filter_status', '按状态过滤任务'),
        ('get_result_success', '获取OCR结果'),
        ('result_no_file', '结果文件不存在'),
        ('retry_success', '重试失败任务'),
        ('retry_not_failed', '非失败状态不可重试'),
        ('delete_success', '删除任务'),
        ('detail_success', '查看任务详情'),
        # SSE stream
        ('publish_progress', '进度事件推送'),
        ('publish_result', '完成事件推送'),
        ('publish_error', '错误事件推送'),
        ('empty_default', '空事件推送'),
        # Table
        ('recognize_basic_table', '基础表格识别'),
        ('simple_table', '简单表格识别'),
        ('recognize_merged_cells', '合并单元格表格识别'),
        ('build_html_table', '表格数据→HTML'),
        ('simple_2x2_table', '2x2表格HTML'),
        ('trim_empty_rows', '空行删除'),
        ('trim_empty_cols', '空列删除'),
        ('remove_empty_row', '单空行删除'),
        ('remove_empty_col', '单空列删除'),
        ('estimate_font_height', '字号估算'),
        ('cluster_1d', '一维聚类(列分隔)'),
        # Text PDF
        ('extract_text_pdf', '文本PDF提取'),
        ('extract_scanned_pdf', '扫描PDF提取(返回空)'),
        ('extract_empty_pdf', '空PDF提取'),
        ('extract_structured_basic', '基础结构化提取'),
        ('extract_structured_skips_empty_lines', '跳过空行'),
        ('global_singleton', '全局单例验证'),
        # Validator
        ('too_many_pages', '超过2000页校验失败'),
        ('default_allowed_extensions', '允许扩展名含音频'),
        ('default_max_pages', 'MAX_PAGES=2000'),
        ('valid_pdf', '正常PDF校验通过'),
        ('encrypted_pdf', '加密PDF校验失败'),
        ('not_pdf_file', '非PDF文件校验失败'),
        ('zero_pages', '0页PDF校验失败'),
        ('negative_page_number', '负数页码校验'),
        ('start_greater_than_end', '起始页>结束页校验'),
        ('end_page_beyond_total', '结束页超出总页数'),
        ('check_text_pdf', '纯文本PDF检测'),
        ('empty_file', '空文件校验'),
        ('empty_pdf_zero_pages', '0页空PDF处理'),
        ('validation_none_code', 'validation_code=None处理'),
        ('no_fitz_fallback', 'fitz不可用时回退'),
        ('no_conflict', '校验无冲突'),
        ('some_pages_empty', '部分页面为空'),
        # Watcher
        ('is_file_stable_true', '文件3次大小不变→已稳定'),
        ('is_file_stable_false', '文件大小变化→未稳定'),
        ('on_created_pdf', '新建PDF事件处理'),
        ('on_created_py', '新建.py文件忽略'),
        ('handle_new_file', '新文件处理流程'),
        ('move_to_error', '无效文件移到error目录'),
        ('move_to_archive', '已完成文件移到archive目录'),
        ('move_to_archive_os_error', '归档OS错误静默处理'),
        ('ensure_dir', '目录不存在自动创建'),
        ('scan_watcher_handler_init', 'ScanWatcherHandler初始化'),
        ('start_watcher', '启动监控服务'),
        ('sets_up_and_starts', '设置并启动所有监控器'),
        ('ensures_all_dirs', '确保所有必要目录存在'),
        # E2E
        ('happy_path', '完整流程:上传→OCR→分析→导出'),
        ('error_recovery', 'OCR失败恢复流程'),
        ('list_filter_sort', '任务列表过滤排序'),
        ('duplicate_detection', '重复上传检测'),
        ('spa_mount', 'SPA前端挂载'),
        ('spa_mount_not_registered_when_dist_missing', 'dist缺失时SPA不注册'),
        ('admin_workflow', '管理员统计+队列'),
        ('auth_chain', '鉴权链路'),
        ('boundary_values', '边界值处理'),
    ]
    
    for pattern, desc in mapping:
        if pattern in n:
            return desc
    
    # Fallback: use the first 40 chars of name, space-separated
    clean = n.replace('_', ' ')
    if len(clean) > 40:
        clean = clean[:40] + '...'
    return clean


def build_test_data(file, cls, name):
    n = name
    specific = [
        ('ping_returns_pong', 'GET /api/v1/ping请求'),
        ('health_when_all', '全部服务正常环境'),
        ('health_when_db', 'PostgreSQL连接断开场景'),
        ('health_when_redis', 'Redis连接断开场景'),
        ('health_when_minio', 'MinIO连接断开场景'),
        ('admin_stats_no', '空数据库(0条任务)'),
        ('admin_stats_return', '含3条任务的数据库'),
        ('admin_queue_return', '含排队任务的数据库'),
        ('admin_queue_empty', '空任务队列'),
        ('too_many_pages', '2001页超大PDF'),
        ('default_allowed', 'PDF+音频扩展名配置'),
        ('default_max_page', 'MAX_PAGES配置值'),
        ('valid_pdf', '5页正常PDF'),
        ('encrypted_pdf', '加密PDF(password=123)'),
        ('not_pdf_file', '.exe非PDF文件'),
        ('identity_keyword', '"原告居民身份证复印件"'),
        ('identity_business', '"被告营业执照副本"'),
        ('identity_credit', '"统一社会信用代码 91310000XXX"'),
        ('multiple_keywords', '"入院病历记录出院小结诊断证明检查报告"'),
        ('all_categories', 'CATEGORY_ORDER完整列表'),
        ('single_fee_item', '1项费用:医疗费12345.67'),
        ('multiple_same_fee', '2项同类:医疗费1000+2000'),
        ('different_fee_type', '2项异类:医疗费5000+交通费800'),
        ('zero_amount', '金额=0的费用项'),
        ('negative_amount', '金额=-100的费用项'),
        ('non_numeric_amount', '金额="N/A"的费用项'),
        ('no_fee_type', '缺fee_type的费用项'),
        ('empty_fee_detail', 'fee_detail={}'),
        ('none_fee_detail', 'fee_detail=None'),
        ('empty_materials', '材料列表=[]'),
        ('rounding', '金额=100.005(需四舍五入)'),
        ('integer_amount', '整数金额:护理费=3000'),
        ('valid_json', 'LLM返回:{"category":"identity_id_card","confidence":0.8}'),
        ('no_json', 'LLM返回非JSON字符串'),
        ('base64_png', '100x100 PNG图片'),
        ('base64_jpeg', '100x100 JPEG图片'),
        ('base64_invalid', '不存在的文件路径'),
        ('factory_bailian', 'ocr_engine_type="bailian"配置'),
        ('factory_default', '默认ocr_engine_type配置'),
        ('real_api', '真实百炼API(需Key)'),
        ('parse_valid_json', '合法JSON数组字符串'),
        ('parse_markdown', 'Markdown代码块包裹的JSON'),
        ('parse_fallback', '纯文本响应'),
        ('parse_empty', '空字符串响应'),
        ('parse_partial', '混杂文本中的JSON'),
        ('continuation_dash', '页尾"—"+页首续接段'),
        ('continuation_ellipsis', '页尾"…"+页首续接段'),
        ('no_continuation', '不同段落跨页'),
        ('simple_merge', '简单跨页续接'),
        ('empty_text', '空字符串/空文本'),
        ('empty_pages', '空页面/空页列表'),
        ('single_page', '单页文档'),
        ('no_merge', '无需合并的文档'),
        ('flat_merge', '无缩进多页段落'),
        ('no_boundary', '无跨页边界'),
        ('callback_success', '有效回调请求体JSON'),
        ('callback_invalid', '无效回调数据'),
        ('callback_timeout', '回调超时场景'),
        ('callback_not_found', '不存在的task_id'),
        ('level_1', '"一、患者基本情况"'),
        ('level_2', '"（一）入院记录"'),
        ('level_3', '"1.主诉"三级标题'),
        ('level_4', '"(1)现病史"四级标题'),
        ('level_5', '"①首次病程"五级标题'),
        ('no_heading', '普通段落文本(无标题特征)'),
        ('mixed_heading', '混合标题+段落文本'),
        ('register_custom', '自定义正则r"^第[一二三四五]部分"'),
        ('registry_contains', 'HeadingPatternsRegistry'),
        ('auto_level', '低对比度灰度图(像素[100,150])'),
        ('deskew', '5度倾斜文档图像'),
        ('crop_black', '含黑边扫描件'),
        ('crop_no_border', '无黑边的正常图像'),
        ('page_numbers_removed', '4页含相同"XX医院"页脚的文档'),
        ('headers_removed', '4页含相同页眉的文档'),
        ('mixed_content_preserved', '4页页眉+正文混合文档'),
        ('text_similarity_identical', '完全相同的两段文本'),
        ('text_similarity_different', '完全不同的两段文本'),
        ('few_pages', '2页文档(少于min_pages阈值)'),
        ('position_is_header', '页面顶部区域文本'),
        ('position_is_footer', '页面底部区域文本'),
        ('upload_small', '1KB文件上传'),
        ('upload_large', '5MB大文件上传'),
        ('upload_empty', '0字节空文件上传'),
        ('download_success', '已存在MinIO对象'),
        ('download_not_found', 'MinIO中不存在的对象'),
        ('delete_success', '已存在MinIO对象删除'),
        ('delete_not_found', '不存在对象删除'),
        ('presigned_get', 'GET预签名URL(3600s)'),
        ('presigned_put', 'PUT预签名URL'),
        ('bucket_exists', 'Bucket存在检查'),
        ('create_bucket', '创建新Bucket'),
        ('ensure_buckets', '自动创建缺失Bucket'),
        ('delete_task_objects', '批量删除task相关对象'),
        ('object_exists', '对象存在检查'),
        ('batch_sequential', '10张图片顺序处理'),
        ('batch_concurrent', '10张图片并发处理(max=3)'),
        ('batch_with_failure', '10张中3张失败'),
        ('batch_empty', '[]空列表批处理'),
        ('empty_image_list', '空图片列表'),
        ('init_cpu', 'CPU模式(enable_mkldnn=False)'),
        ('load_model', 'PaddleOCR模型预加载'),
        ('recognize_single', '单张图片OCR'),
        ('recognize_batch', '5张图片批量OCR'),
        ('recognize_not_loaded', '模型未加载状态'),
        ('recognize_empty_page', '空页面OCR'),
        ('save_result', 'OCR结果JSON文件'),
        ('detect_0_degrees', '0度正常方向页面'),
        ('detect_90_degrees', '90度旋转页面'),
        ('detect_180_degrees', '180度旋转页面'),
        ('detect_270_degrees', '270度旋转页面'),
        ('split_paragraphs', '3段连续文本'),
        ('group_paragraphs', '含缩进层级5段文本'),
        ('build_hierarchy', '3级缩进层级结构'),
        ('page_number', '页码纯数字文本'),
        ('classify_text_pdf', '纯文本PDF特征'),
        ('classify_scanned', '扫描件PDF特征'),
        ('classify_empty', '空PDF(0页)'),
        ('pdf_info', 'PDF元信息(页数/加密/大小)'),
        ('default_dpi', '5页PDF+DPI=200'),
        ('split_all_pages', '5页PDF全拆'),
        ('split_page_range', '5页PDF拆第2-4页'),
        ('end_page_beyond', '5页PDF end_page=100'),
        ('output_filenames', '5页PDF输出文件名格式'),
        ('dpi_zoom', '5页PDF DPI=150缩放矩阵'),
        ('split_to_bytes', '字节流拆分'),
        ('ocr_confidence', 'OCR置信度0.85结果'),
        ('heading_quality', '5级标题结构'),
        ('max_heading_depth', '最大深度=3'),
        ('structure_completeness', '完整结构(标题+段落+表格)'),
        ('score_structure', '基础评分计算'),
        ('empty_ocr', '空OCR结果({})'),
        ('export_basic', '3段落OCR结果'),
        ('export_with_table', '含1个表格OCR结果'),
        ('export_empty', '空导出结果'),
        ('happy_path', '完整E2E管线:上传→OCR→分析→导出'),
        ('error_recovery', 'OCR失败→重试→成功流程'),
        ('list_filter_sort', '任务列表过滤+排序'),
        ('duplicate_detection', '相同PDF二次上传'),
        ('spa_mount', '前端SPA静态资源'),
        ('admin_workflow', '管理员统计+队列API'),
        ('auth_chain', 'X-API-Key鉴权流程'),
    ]
    for pattern, data in specific:
        if pattern in n:
            return data
    return '单元测试Mock数据'


# Build all rows
rows = []
for test in tests:
    file_ = test['file']
    cls = test['class'] or ''
    name = test['name']
    status = test['status']
    
    func_name = module_map.get(file_, file_.replace('test_', '').replace('.py', ''))
    if cls:
        func_name = f"{func_name} \u2192 {cls}"
    
    # Use docstring as expected result if available
    doc = docstrings.get(name, '')
    expected = doc if doc else '断言通过,结果符合预期'
    if len(expected) > 100:
        expected = expected[:97] + '...'
    
    actual = '与预计结果一致,断言通过' if status == 'PASSED' else '跳过(需百炼API Key)' if status == 'SKIPPED' else '失败'
    pass_val = '\u2705 通过' if status == 'PASSED' else '\u23ed 跳过' if status == 'SKIPPED' else '\u274c 失败'
    
    rows.append({
        '功能名称': func_name,
        '测试数据': build_test_data(file_, cls, name),
        '操作步骤': humanize_test_name(name),
        '预计结果': expected,
        '实际执行结果': actual,
        '是否通过': pass_val,
        '测试人员': '自动化测试',
        '开发人员': 'ScanStruct团队',
        '复核人员': '\u2014',
    })

df = pd.DataFrame(rows)

# Dedup
content_cols = ['功能名称', '测试数据', '操作步骤', '预计结果']
before = len(df)
df = df.drop_duplicates(subset=content_cols, keep='first').reset_index(drop=True)
after = len(df)
print(f'Dedup: {before} → {after} (removed {before - after})')

df.insert(0, '案例序号', range(1, len(df) + 1))

# Create Excel
wb = Workbook()
ws = wb.active
ws.title = "测试用例报告"

header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2F5496')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_font = Font(name='Arial', size=10)
cell_align = Alignment(vertical='center', wrap_text=True)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
pass_fill = PatternFill('solid', fgColor='C6EFCE')
skip_fill = PatternFill('solid', fgColor='FFEB9C')
pass_font = Font(name='Arial', size=10, color='006100')
skip_font = Font(name='Arial', size=10, color='9C5700')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

headers_list = ['案例序号', '功能名称', '测试数据', '操作步骤', '预计结果', '实际执行结果', '是否通过', '测试人员', '开发人员', '复核人员']
for col, h in enumerate(headers_list, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = header_align
    c.border = thin_border

for idx, row_data in df.iterrows():
    row = idx + 2
    for col, h in enumerate(headers_list, 1):
        val = row_data[h]
        c = ws.cell(row=row, column=col, value=val)
        c.font = cell_font
        c.alignment = center_align if col in (1, 7, 8, 9, 10) else cell_align
        c.border = thin_border
        if col == 7:
            if '\u2705' in str(val):
                c.fill = pass_fill; c.font = pass_font
            elif '\u23ed' in str(val):
                c.fill = skip_fill; c.font = skip_font

col_widths = [10, 34, 26, 28, 34, 22, 12, 14, 16, 12]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:J{after+1}'

# Sheet 2
ws2 = wb.create_sheet("统计汇总")
mod_stats = defaultdict(lambda: {'total': 0, 'passed': 0, 'skipped': 0})
for _, r in df.iterrows():
    mod = r['功能名称'].split(' \u2192 ')[0]
    mod_stats[mod]['total'] += 1
    if '\u2705' in str(r['是否通过']): mod_stats[mod]['passed'] += 1
    elif '\u23ed' in str(r['是否通过']): mod_stats[mod]['skipped'] += 1

sh = ['功能模块', '去重后用例', '通过', '跳过', '通过率']
for col, h in enumerate(sh, 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = header_align; c.border = thin_border

ri = 2
for mod in sorted(mod_stats, key=lambda x: mod_stats[x]['total'], reverse=True):
    s = mod_stats[mod]
    rate = f"{s['passed']/s['total']*100:.1f}%"
    for col, val in enumerate([mod, s['total'], s['passed'], s['skipped'], rate], 1):
        c = ws2.cell(row=ri, column=col, value=val)
        c.font = cell_font; c.alignment = center_align if col >= 2 else cell_align; c.border = thin_border
    ri += 1

tp = sum(1 for _, r in df.iterrows() if '\u2705' in str(r['是否通过']))
ts = sum(1 for _, r in df.iterrows() if '\u23ed' in str(r['是否通过']))
for col, val in enumerate(['合计', after, tp, ts, f"{tp/after*100:.1f}%"], 1):
    c = ws2.cell(row=ri, column=col, value=val)
    c.font = Font(name='Arial', size=10, bold=True); c.alignment = center_align if col >= 2 else cell_align
    c.border = thin_border; c.fill = PatternFill('solid', fgColor='D9E2F3')

ws2.column_dimensions['A'].width = 28
for cl in ['B', 'C', 'D', 'E']:
    ws2.column_dimensions[cl].width = 14
ws2.freeze_panes = 'A2'

output = 'C:/Users/Administrator/Desktop/test_report_2026-06-06.xlsx'
wb.save(output)
print(f'Saved: {output}')
print(f'Sheet1: {after} rows after dedup')
