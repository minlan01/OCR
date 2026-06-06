"""生成深度用户实测 Excel 报告"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open("docs/api_test_results_2026-06-06.json", "r", encoding="utf-8") as f:
    results = json.load(f)

wb = Workbook()
ws = wb.active
ws.title = "深度用户实测报告"

# Headers
headers = ["案例序号", "功能名称", "测试数据", "操作步骤", "预计结果", "实际执行结果", "是否通过", "测试人员", "开发人员", "复核人员"]

# Styles
hdr_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="2F5496")
hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_font = Font(name="Microsoft YaHei", size=10)
cell_align = Alignment(vertical="top", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
pass_fill = PatternFill("solid", fgColor="C6EFCE")
fail_fill = PatternFill("solid", fgColor="FFC7CE")
pass_font = Font(name="Microsoft YaHei", size=10, color="006100", bold=True)
fail_font = Font(name="Microsoft YaHei", size=10, color="9C0006", bold=True)
bug_fill = PatternFill("solid", fgColor="FFEB9C")
bug_font = Font(name="Microsoft YaHei", size=10, color="9C5700", bold=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

# Write headers
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = hdr_align
    c.border = thin_border

# Write data
for idx, r in enumerate(results):
    row = idx + 2
    for col, key in enumerate(headers, 1):
        val = r.get(key, "")
        c = ws.cell(row=row, column=col, value=val)
        c.font = cell_font
        c.alignment = cell_align if col in (3,4,5,6) else center_align
        c.border = thin_border
        
        # "是否通过"列特殊样式
        if col == 7:
            c.alignment = center_align
            if "✅" in str(val):
                c.fill = pass_fill
                c.font = pass_font
            elif "❌" in str(val):
                is_bug = "BUG" in str(r.get("功能名称", ""))
                if is_bug:
                    c.fill = bug_fill
                    c.font = bug_font
                    c.value = "⚠ Bug"
                else:
                    c.fill = fail_fill
                    c.font = fail_font

# Column widths
widths = [8, 22, 36, 34, 34, 42, 10, 12, 12, 8]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Freeze + auto-filter
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:J{len(results)+1}"

# Row height for header
ws.row_dimensions[1].height = 28

# ═══ Sheet 2: 统计汇总 ═══
ws2 = wb.create_sheet("统计汇总")

# Category stats
from collections import defaultdict
cat_stats = defaultdict(lambda: {"total": 0, "pass": 0, "fail": 0, "bug": 0})
for r in results:
    func = r["功能名称"]
    # Determine category
    if "系统" in func or "健康" in func or "连通" in func or "队列" in func or "统计" in func:
        cat = "基础设施"
    elif "案件" in func and "创建" in func:
        cat = "案件创建"
    elif "案件" in func and ("获取" in func or "列表" in func):
        cat = "案件查询"
    elif "案件" in func and ("更新" in func or "删除" in func):
        cat = "案件修改/删除"
    elif any(k in func for k in ["死亡诊断", "日期提取", "事实", "结论", "鉴定", "时间线", "检查项", "被告"]):
        cat = "分析结果验证"
    elif "目录" in func or "验证" in func or "费用" in func:
        cat = "目录/验证"
    elif "导出" in func or "打包" in func or "PDF" in func:
        cat = "导出功能"
    elif "进度" in func or "处理" in func:
        cat = "流程控制"
    elif "扫描" in func:
        cat = "扫描系统"
    elif "模板" in func:
        cat = "模板系统"
    else:
        cat = "其他"
    
    cat_stats[cat]["total"] += 1
    if "✅" in r["是否通过"]:
        cat_stats[cat]["pass"] += 1
    elif "Bug" in r.get("功能名称", ""):
        cat_stats[cat]["bug"] += 1
        cat_stats[cat]["fail"] += 1
    else:
        cat_stats[cat]["fail"] += 1

summary_headers = ["功能模块", "总用例", "通过", "失败", "Bug数", "通过率"]
for col, h in enumerate(summary_headers, 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = hdr_align
    c.border = thin_border

row_idx = 2
for cat in sorted(cat_stats.keys(), key=lambda x: cat_stats[x]["total"], reverse=True):
    s = cat_stats[cat]
    rate = f"{s['pass']/s['total']*100:.1f}%"
    data = [cat, s["total"], s["pass"], s["fail"], s.get("bug", 0), rate]
    for col, val in enumerate(data, 1):
        c = ws2.cell(row=row_idx, column=col, value=val)
        c.font = cell_font
        c.alignment = center_align if col >= 2 else cell_align
        c.border = thin_border
        if col == 6 and rate == "100.0%":
            c.fill = pass_fill
            c.font = pass_font
    row_idx += 1

# Total
total_pass = sum(1 for r in results if "✅" in r["是否通过"])
total_fail = sum(1 for r in results if "❌" in r["是否通过"])
total_bug = sum(1 for r in results if "Bug" in r.get("功能名称", ""))
total = len(results)
total_data = ["合计", total, total_pass, total_fail, total_bug, f"{total_pass/total*100:.1f}%"]
for col, val in enumerate(total_data, 1):
    c = ws2.cell(row=row_idx, column=col, value=val)
    c.font = Font(name="Microsoft YaHei", size=10, bold=True)
    c.alignment = center_align if col >= 2 else cell_align
    c.border = thin_border
    c.fill = PatternFill("solid", fgColor="D9E2F3")

ws2.column_dimensions["A"].width = 18
for ch in "BCDEF":
    ws2.column_dimensions[ch].width = 12
ws2.freeze_panes = "A2"

# ═══ Sheet 3: Bug清单 ═══
ws3 = wb.create_sheet("Bug清单")
bug_headers = ["序号", "Bug描述", "触发端点", "根因", "严重级别", "修复优先级"]
for col, h in enumerate(bug_headers, 1):
    c = ws3.cell(row=1, column=col, value=h)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = hdr_align
    c.border = thin_border

bugs = [
    (1, "导出立案证据 500 → 已修复", "GET /evidence/cases/{id}/export/filing-evidence",
     "原: Cannot run the event loop while another loop is running — 改用 _inline_data 函数直接传数据,避免 run_in_worker",
     "P1 高", "已修复"),
    (2, "导出赔偿费用清单 500 → 已修复", "GET /evidence/cases/{id}/export/compensation",
     "同 Bug#1: 改用 generate_compensation_inline_data",
     "P1 高", "已修复"),
    (3, "导出司法鉴定申请书 500 → 已修复", "GET /evidence/cases/{id}/export/appraisal-app",
     "同 Bug#1: 改用 generate_appraisal_inline_data",
     "P1 高", "已修复"),
    (4, "目录PDF生成 500 → 已修复", "GET /evidence/cases/{id}/catalog/pdf",
     "原: 'ChineseFont' KeyError — Docker容器内置 simhei.ttf + 修复 _ensure_chinese_font 全局变量逻辑",
     "P1 高", "已修复"),
    (5, "材料PDF列表 500 → 已修复", "GET /evidence/cases/{id}/materials/pdf",
     "同 Bug#4: ChineseFont 字体缺失, 同修复方案",
     "P2 中", "已修复"),
]

for idx, (seq, desc, ep, cause, level, pri) in enumerate(bugs):
    row = idx + 2
    data = [seq, desc, ep, cause, level, pri]
    for col, val in enumerate(data, 1):
        c = ws3.cell(row=row, column=col, value=val)
        c.font = cell_font
        c.alignment = cell_align if col >= 2 else center_align
        c.border = thin_border

ws3.column_dimensions["A"].width = 8
ws3.column_dimensions["B"].width = 28
ws3.column_dimensions["C"].width = 45
ws3.column_dimensions["D"].width = 55
ws3.column_dimensions["E"].width = 10
ws3.column_dimensions["F"].width = 10
ws3.freeze_panes = "A2"

# Save
output = "C:/Users/Administrator/Desktop/test_report_api_2026-06-06.xlsx"
wb.save(output)
print(f"已保存至 {output}")
print(f"共 {len(results)} 条测试, {total_pass} 通过, {total_fail} 失败(含 {total_bug} Bug)")
