"""Generate test report Excel from test_dataset_2026-06-06.json"""
import json
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open('docs/test_dataset_2026-06-06.json', 'r', encoding='utf-8') as f:
    tests = json.load(f)

module_map = {
    'test_api.py': 'API 健康检查与管理',
    'test_bailian_ocr.py': 'OCR 引擎(百炼)',
    'test_callback.py': '业务回调',
    'test_cross_page_merger.py': '跨页段落合并',
    'test_e2e_workflows.py': '端到端工作流',
    'test_evidence.py': '证据分类/目录/分析/导出',
    'test_header_footer_cleaner.py': '页眉页脚清除',
    'test_heading_parser.py': '标题层级解析',
    'test_image_enhancer.py': '图像增强',
    'test_json_exporter.py': 'JSON 导出',
    'test_layout_detector.py': '版面检测',
    'test_list_detector.py': '列表检测',
    'test_minio_client.py': 'MinIO 存储客户端',
    'test_ocr_batch_processor.py': 'OCR 批处理',
    'test_ocr_engine.py': 'OCR 引擎(PaddleOCR)',
    'test_orientation_standalone.py': '页面方向检测',
    'test_paragraph_grouper.py': '段落分组',
    'test_pdf_classifier.py': 'PDF 分类器',
    'test_pdf_splitter.py': 'PDF 页面拆分',
    'test_quality_scorer.py': '质量评分',
    'test_scan_api.py': '扫描 API',
    'test_stream_publisher.py': 'SSE 流式推送',
    'test_table_recognizer.py': '表格识别',
    'test_text_pdf_extractor.py': '文本 PDF 提取',
    'test_validator.py': 'PDF 校验器',
    'test_watcher.py': '文件监控',
}


def infer_test_data(file, cls, name):
    kw = name.lower()
    if 'encrypt' in kw: return '加密PDF文件'
    if 'orientation' in kw or 'rotat' in kw: return '旋转/方向异常的PDF'
    if 'blank' in kw or 'empty' in kw: return '空白/空PDF文件'
    if 'large' in kw or 'big' in kw: return '大文件/超长PDF'
    if 'split' in kw: return '多页PDF文件'
    if 'fee' in kw or 'cost' in kw: return '费用明细数据(医疗费/护理费等)'
    if 'identity' in kw or 'card' in kw: return '身份证/营业执照等身份材料'
    if 'death' in kw: return '死亡案件相关材料'
    if 'injury' in kw: return '伤残案件相关材料'
    if 'neonatal' in kw: return '新生儿案件相关材料'
    if ('ocr' in kw and 'import' not in kw): return 'OCR识别测试图片/PDF'
    if 'image' in kw and 'enhanc' in kw: return '待增强图像(低对比度/模糊/倾斜)'
    if 'minio' in file or 'upload' in kw or 'download' in kw: return 'MinIO文件上传/下载对象'
    if 'table' in kw: return '含表格的PDF/图片'
    if 'header' in kw or 'footer' in kw: return '含页眉页脚的PDF'
    if 'heading' in kw: return '含标题层级的文档'
    if 'list_det' in kw: return '含列表结构的文档'
    if 'paragraph' in kw or 'grouper' in kw: return '含多段落的文档'
    if 'cross_page' in kw: return '跨页连续段落文档'
    if 'quality' in kw: return '质量评分样本图片'
    if 'watch' in kw or 'monitor' in kw: return '监控目录文件事件'
    if 'callback' in kw: return '回调请求体JSON'
    if 'stream' in kw or 'sse' in kw: return 'SSE流式数据'
    if 'export' in kw or 'json' in kw: return '导出/结构化数据对象'
    if 'catalog' in kw: return '证据目录数据'
    if 'health' in kw: return 'API健康检查请求'
    if 'ping' in kw: return 'Ping请求'
    if 'schema' in kw or 'response' in kw or 'request' in kw: return 'Pydantic Schema定义'
    if 'endpoint' in kw or 'route' in kw: return 'API路由定义'
    if 'llm' in kw: return 'LLM分类/提取请求'
    if 'classify' in kw: return '证据分类输入文本'
    if 'title' in kw or 'purpose' in kw: return '分类标题/用途生成输入'
    if 'diagn' in kw: return '诊断标签提取数据'
    if 'date' in kw: return '日期格式化字符串'
    if 'normaliz' in kw: return '待标准化数据'
    if 'import' in kw: return '模块导入验证'
    if 'init' in kw: return '模块初始化验证'
    if 'validator' in file: return 'PDF文件(页数/格式/加密状态)'
    if 'enhanc' in file: return '图像处理测试数据'
    if 'layout' in file: return '版面检测结果对象'
    if 'stream' in file: return 'SSE流配置'
    if 'scan' in file: return '扫描API请求/响应'
    if 'text_pdf' in file: return '文本型PDF文件'
    if 'pdf_class' in file: return 'PDF分类器输入'
    return '单元测试Mock数据'


def infer_steps(file, cls, name):
    kw = name.lower()
    parts = []
    parts.append('1. 准备测试数据' if 'mock' not in kw else '1. 构造Mock对象/数据')
    if 'encrypt' in kw: step = '2. 上传加密PDF进行校验'
    elif 'split' in kw: step = '2. 调用PDF拆分方法split_to_images()'
    elif 'classify' in kw: step = '2. 调用分类器classify_text()'
    elif 'enhanc' in kw: step = '2. 调用图像增强处理'
    elif 'ocr' in kw and 'bailian' in file: step = '2. 调用百炼OCR API识别'
    elif 'ocr' in kw: step = '2. 执行OCR识别'
    elif 'upload' in kw: step = '2. 调用MinIO上传文件'
    elif 'download' in kw: step = '2. 调用MinIO下载文件'
    elif 'health' in kw: step = '2. GET /api/v1/health'
    elif 'ping' in kw: step = '2. GET /api/v1/ping'
    elif 'export' in kw: step = '2. 执行导出操作'
    elif 'fee' in kw: step = '2. 调用_calculate_fee_summary()'
    elif 'diagn' in kw: step = '2. 调用诊断标签提取函数'
    elif 'date' in kw or 'normaliz' in kw: step = '2. 调用normalize_date()'
    elif 'validate' in kw or 'validator' in file: step = '2. 调用PDF校验器validate()'
    elif 'watch' in kw or 'monitor' in kw: step = '2. 触发文件系统事件'
    elif 'stream' in kw: step = '2. 创建SSE流并推送数据'
    elif 'table' in kw: step = '2. 调用表格识别'
    elif 'heading' in kw: step = '2. 调用标题解析器'
    elif 'header' in kw or 'footer' in kw: step = '2. 调用clean_headers_footers()'
    elif 'import' in kw: step = '2. 执行import语句'
    elif 'init' in kw: step = '2. 执行模块初始化'
    elif 'endpoint' in kw: step = '2. 遍历FastAPI路由表'
    elif 'schema' in kw or 'response' in kw: step = '2. 比对Schema字段定义'
    elif 'llm' in kw: step = '2. 调用LLM分类接口'
    elif 'title' in kw: step = '2. 调用_generate_title()'
    elif 'purpose' in kw: step = '2. 调用_generate_proof_purpose()'
    elif 'merge' in kw: step = '2. 调用跨页合并函数'
    elif 'group' in kw: step = '2. 调用段落分组函数'
    elif 'quality' in kw: step = '2. 调用质量评分函数'
    elif 'detect' in kw: step = '2. 调用检测函数'
    elif 'clean' in kw: step = '2. 调用清洗函数'
    elif 'move' in kw or 'archive' in kw: step = '2. 触发文件移动/归档'
    elif 'scan' in file and 'api' in file: step = '2. 调用扫描API端点'
    elif 'text_pdf' in file: step = '2. 调用文本PDF提取'
    else: step = '2. 调用目标方法/函数'
    parts.append(step)
    parts.append('3. 断言返回结果符合预期')
    return '\n'.join(parts)


def infer_expected(name):
    kw = name.lower()
    mapping = [
        ('encrypt', '返回加密检测结果(encrypted=True/False)'),
        ('orientation', '返回正确的页面旋转角度'),
        ('split', '返回正确数量的页面图片'),
        ('identity_id_card', '分类结果为identity_id_card'),
        ('identity_defendant', '分类结果为identity_defendant'),
        ('classify', '返回正确的分类及置信度'),
        ('fee', '返回正确的费用汇总金额'),
        ('health', '返回健康状态JSON'),
        ('ping', '返回{"ping":"pong"}'),
        ('ocr', '返回OCR识别文本结果'),
        ('upload', '文件上传成功,返回对象路径'),
        ('download', '文件下载成功,返回文件内容'),
        ('enhanc', '图像增强后品质提升/参数正确'),
        ('table', '返回识别的表格结构'),
        ('heading', '返回正确的标题层级'),
        ('validate', '校验结果:valid/invalid+错误信息'),
        ('too_many', '校验结果:invalid+超页数错误'),
        ('diagn', '返回匹配的诊断标签列表'),
        ('date', '返回标准化日期格式(YYYY-MM-DD)'),
        ('normaliz', '返回标准化结果'),
        ('export', '成功导出目标格式文件'),
        ('schema', 'Schema字段完整无缺失'),
        ('response', 'Response Schema字段覆盖Model'),
        ('endpoint', '端点数量/路由注册符合预期'),
        ('callback', '回调处理返回正确结果'),
        ('stream', 'SSE流正确推送事件'),
        ('title', '生成格式正确的分类标题'),
        ('purpose', '生成正确的证明用途描述'),
        ('llm', 'LLM返回有效JSON并正确解析'),
        ('import', '模块可正常导入,无异常'),
        ('init', '模块初始化正常'),
        ('merge', '跨页段落正确合并'),
        ('group', '段落正确分组'),
        ('quality', '返回正确的质量评分'),
        ('detect', '检测结果符合预期'),
        ('clean', '清洗结果符合预期'),
        ('move', '文件正确归档/移动'),
        ('archive', '文件正确归档'),
    ]
    for key, val in mapping:
        if key in kw:
            return val
    return '测试断言全部通过,无异常'


# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "测试用例报告"

headers = ['案例序号', '功能名称', '测试数据', '操作步骤', '预计结果', '实际执行结果', '是否通过', '测试人员', '开发人员', '复核人员']

# Styles
header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2F5496')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_font = Font(name='Arial', size=10)
cell_align = Alignment(vertical='center', wrap_text=True)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
pass_fill = PatternFill('solid', fgColor='C6EFCE')
skip_fill = PatternFill('solid', fgColor='FFEB9C')
fail_fill = PatternFill('solid', fgColor='FFC7CE')
pass_font = Font(name='Arial', size=10, color='006100')
skip_font = Font(name='Arial', size=10, color='9C5700')
fail_font = Font(name='Arial', size=10, color='9C0006')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Write headers
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# Write data
for idx, test in enumerate(tests, 1):
    row = idx + 1
    file_ = test['file']
    cls = test['class'] or ''
    name = test['name']
    status = test['status']

    ws.cell(row=row, column=1, value=idx).alignment = center_align
    func_name = module_map.get(file_, file_.replace('test_', '').replace('.py', ''))
    if cls:
        func_name = f"{func_name} \u2192 {cls}"
    ws.cell(row=row, column=2, value=func_name).alignment = cell_align
    ws.cell(row=row, column=3, value=infer_test_data(file_, cls, name)).alignment = cell_align
    ws.cell(row=row, column=4, value=infer_steps(file_, cls, name)).alignment = cell_align
    ws.cell(row=row, column=5, value=infer_expected(name)).alignment = cell_align

    if status == 'PASSED':
        actual = '与预计结果一致,断言全部通过'
    elif status == 'SKIPPED':
        actual = '跳过(需百炼API Key或条件不满足)'
    else:
        actual = '与预计结果不一致'
    ws.cell(row=row, column=6, value=actual).alignment = cell_align

    if status == 'PASSED':
        pass_val, fill, font = '\u2705 通过', pass_fill, pass_font
    elif status == 'SKIPPED':
        pass_val, fill, font = '\u23ed 跳过', skip_fill, skip_font
    else:
        pass_val, fill, font = '\u274c 失败', fail_fill, fail_font
    c7 = ws.cell(row=row, column=7, value=pass_val)
    c7.alignment = center_align
    c7.fill = fill
    c7.font = font

    ws.cell(row=row, column=8, value='自动化测试').alignment = center_align
    ws.cell(row=row, column=9, value='ScanStruct团队').alignment = center_align
    ws.cell(row=row, column=10, value='\u2014').alignment = center_align

    for col in range(1, 11):
        c = ws.cell(row=row, column=col)
        c.border = thin_border
        if col != 7:
            c.font = cell_font

col_widths = [10, 32, 24, 30, 32, 26, 12, 14, 16, 12]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:J{len(tests)+1}'

# ===== Sheet 2 =====
ws2 = wb.create_sheet("统计汇总")
module_stats = defaultdict(lambda: {'total': 0, 'passed': 0, 'skipped': 0, 'failed': 0})
for t in tests:
    f = t['file']
    module_stats[f]['total'] += 1
    s = t['status'].lower()
    if s == 'passed': module_stats[f]['passed'] += 1
    elif s == 'skipped': module_stats[f]['skipped'] += 1
    else: module_stats[f]['failed'] += 1

summary_headers = ['模块文件', '功能名称', '总用例', '通过', '跳过', '失败', '通过率']
for col, h in enumerate(summary_headers, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

row_idx = 2
for f in sorted(module_stats.keys(), key=lambda x: module_stats[x]['total'], reverse=True):
    s = module_stats[f]
    rate = f"{s['passed']/s['total']*100:.1f}%"
    desc = module_map.get(f, f.replace('test_', '').replace('.py', ''))
    data = [f, desc, s['total'], s['passed'], s['skipped'], s['failed'], rate]
    for col, val in enumerate(data, 1):
        cell = ws2.cell(row=row_idx, column=col, value=val)
        cell.font = cell_font
        cell.alignment = center_align if col >= 3 else cell_align
        cell.border = thin_border
        if col == 7 and rate == '100.0%':
            cell.fill = pass_fill
            cell.font = pass_font
    row_idx += 1

total_data = ['合计', '', len(tests),
              sum(1 for t in tests if t['status'] == 'PASSED'),
              sum(1 for t in tests if t['status'] == 'SKIPPED'),
              sum(1 for t in tests if t['status'] == 'FAILED'),
              f"{sum(1 for t in tests if t['status'] == 'PASSED') / len(tests) * 100:.1f}%"]
for col, val in enumerate(total_data, 1):
    cell = ws2.cell(row=row_idx, column=col, value=val)
    cell.font = Font(name='Arial', size=10, bold=True)
    cell.alignment = center_align if col >= 3 else cell_align
    cell.border = thin_border
    cell.fill = PatternFill('solid', fgColor='D9E2F3')

ws2.column_dimensions['A'].width = 28
ws2.column_dimensions['B'].width = 24
for c in ['C', 'D', 'E', 'F', 'G']:
    ws2.column_dimensions[c].width = 12
ws2.freeze_panes = 'A2'

output_path = 'docs/test_report_2026-06-06.xlsx'
wb.save(output_path)
print(f'Saved to {output_path}')
print(f'Total rows: {len(tests)}')
print(f'Sheets: {wb.sheetnames}')
